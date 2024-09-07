from openpyxl import load_workbook, Workbook 
from openpyxl.styles import Font, Border, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os


def copy_style(source_cell, target_cell):
    """셀의 스타일을 복사하는 함수"""
    if source_cell.font:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            vertAlign=source_cell.font.vertAlign,
            underline=source_cell.font.underline,
            strike=source_cell.font.strike,
            color=source_cell.font.color
        )

    if source_cell.border:
        target_cell.border = Border(
            left=source_cell.border.left,
            right=source_cell.border.right,
            top=source_cell.border.top,
            bottom=source_cell.border.bottom,
            diagonal=source_cell.border.diagonal,
            diagonal_direction=source_cell.border.diagonal_direction,
            outline=source_cell.border.outline,
            vertical=source_cell.border.vertical,
            horizontal=source_cell.border.horizontal
        )

    if source_cell.alignment:
        target_cell.alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            text_rotation=source_cell.alignment.text_rotation,
            wrap_text=source_cell.alignment.wrap_text,
            shrink_to_fit=source_cell.alignment.shrink_to_fit,
            indent=source_cell.alignment.indent
        )

def copy_row_height_and_column_width(table_widget, template_sheet:Workbook.worksheets, target_sheet:Workbook.worksheets):
    """행 높이와 열 너비를 복사하는 함수"""
    for row in range(table_widget.rowCount()):
        for template_row in range(1, template_sheet.max_row + 1):
            new_row = row * template_sheet.max_row + template_row
            target_sheet.row_dimensions[new_row].height = template_sheet.row_dimensions[template_row].height

    for col in range(1, template_sheet.max_column + 1):
        target_sheet.column_dimensions[get_column_letter(col)].width = template_sheet.column_dimensions[get_column_letter(col)].width

def apply_merged_cells(template_sheet, report_sheet, row_count):
    """병합된 셀을 템플릿 시트에서 레포트 시트에 반복적으로 적용"""
    for merged_range in template_sheet.merged_cells.ranges:
        # 병합된 셀의 시작과 끝 좌표
        min_row, min_col, max_row, max_col = merged_range.bounds
        
        for r in range(row_count):
            # 병합된 셀의 범위를 조정하여 레포트 시트에 적용
            new_min_row = min_row + r * row_count
            new_max_row = max_row + r * row_count
            report_sheet.merge_cells(start_row=new_min_row, start_column=min_col, end_row=new_max_row, end_column=max_col)

def extract_column_headers(table_widget) -> list:
    """
    테이블 위젯의 헤더를 리스트로 반환
    """
    headers = []
    for column in range(table_widget.columnCount()):
        item = table_widget.horizontalHeaderItem(column)
        if item is not None:
            headers.append(item.text())
    return headers

def export_to_report(table_widget, template_path, report_path):
    """
    테이블 위젯의 데이터를 레포트 파일(xlsx)로 저장
    param table_widget:(QTableWidget), 실제 data가 기록된 테이블위젯
    param template_path:(str), 서식 파일(xlsx)
    param report_path: (str), 저장할 파일(xlsx)  
    """

    # 템플릿 파일 로드
    template_wb = load_workbook(template_path)
    template_sheet = template_wb['서식']

    # 리포트 파일 생성
    report_wb = Workbook()
    report_sheet = report_wb.create_sheet('report', 0)

    # 템플릿 시트를 row 수만큼 복사
    for row in range(table_widget.rowCount()):
        for template_row in range(1, template_sheet.max_row + 1):  # 템플릿의 모든 행
            new_row = row * template_sheet.max_row + template_row
            for column in range(1, template_sheet.max_column + 1):
                # 템플릿에서 데이터 가져오기
                template_cell = template_sheet.cell(row=template_row, column=column)
                report_cell = report_sheet.cell(row=new_row, column=column)

                # 데이터 복사
                report_cell.value = template_cell.value

                # 스타일 복사
                copy_style(template_cell, report_cell)
    # 서식의 행높이, 열너비 복사
    copy_row_height_and_column_width(table_widget, template_sheet, report_sheet)
    # 병합된 셀 적용
    apply_merged_cells(template_sheet, report_sheet, table_widget.rowCount())

    ##### QTableWidget의 데이터 입력 #####
    # 1. 데이터의 개별값이 단순처리되어 입력되는 경우 좌표 지정
    # template_coordinate={"점번호":           {"coord":(4, 2)},
    #                     "X":               {"coord":(9, 2)},
    #                     "Y":               {"coord":(9, 5)},
    #                     "도선등급":         {"coord":(3, 7)},
    #                     "도선명":           {"coord": (3,12)},
    #                     "표지재질":         {"coord":(3, 26)},
    #                     # "토지소재(동리)":   {"coord":(5, 2), "callback":None}, ##### 토지소재란에 동리 + 지번
    #                     # "토지소재(지번)":   {"coord":(5, 2), "callback":None}, #####  
    #                     "지적도도호":       {"coord":(5, 26)}, 
    #                     "설치년월일":       {"coord":(8, 1), "callback":None},
    #                     "조사년월일":       {"coord":(21, 1), "callback":None},                        
    #                     "조사자(직)":       {"coord":(21, 2)},
    #                     "조사자(성명)":     {"coord":(21, 5)},
    #                     "경위도(L)":        {"coord":(9, 11), "callback":None},
    #                     "경위도(B)":        {"coord":(9, 22), "callback":None}
    # }

    # for row in range(table_widget.rowCount()):
    #     data = [table_widget.item(row, column).text() if table_widget.item(row, column) is not None else "" for column in range(table_widget.columnCount())]
    #     headers = extract_column_headers(table_widget)
    #     data_dic = {h:v for h, v in zip(headers, data)}
    #     # 각 값들에 대한 템플릿의 위치 지정
    #     for key in template_coordinate:
    #         target_row = template_coordinate[key].get("coord")[0]
    #         target_col = template_coordinate[key].get("coord")[1]
    #         value = data_dic[key] if template_coordinate[key].get("callback", None) is None else template_coordinate[key].get("callback")(data_dic[key])

    #         report_sheet.cell(row=row * template_sheet.max_row + target_row, column=column + target_col, value=value)
    
        # # 2. 데이터 값이 서로 더해지거나 사진입력의 경우 
        # loc = ' '.join(data_dic["토지소재(동리)"], data_dic["토지소재(지번)"])
        # img_file = os.path.join(data_dic["사진파일(경로)"], data_dic["사진파일명"])

        # # 토지소재 = 토지소재(동리) + 지번
        # report_sheet.cell(row=row * template_sheet.max_row + 5, column=column + 2, value=loc)
        # # 이미지 파일 입력
        # pass



    ##### 데이터 입력 완료 ######

    # 리포트 파일 저장
    report_wb.save(report_path)
    print(f"리포트 파일로 내보내기 완료: {report_path}")