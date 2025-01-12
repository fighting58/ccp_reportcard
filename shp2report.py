import openpyxl_addin as oa
import openpyxl
import pandas as pd
from copy import copy


class ReportFromDataframe():
    """
    DataFrame 형식자료를 입력받아 서식(template, xlsx)에 자료를 입력하여 보고서 작성
    """

    def __init__(self, template: str, sheetname:str = None, savefile:str = None, dataframe: pd.DataFrame=None, 
                 max_row: int = None, border_settings:list = None, mappings:list = None) -> None:
        """
        클래스 초기화
        params template: str, 서식 파일(.xlsx)
               sheetname: str, 서식파일의 시트명, 지정하지 않으면 처음 시트
               savefile:str, 저장할 파일(.xlsx)
               dataframe: pandas.DataFrame, 입력 데이터
               max_row: 서식 한 페이지의 최대 row 개수, 미지정시 template의 마지막 열
               border_settings: list(dict), 서식의 경계선 정의
               mappings: list(dict), 데이터 필드 및 좌표 지정, 처리함수의 dictionary
        """
        self.template = template
        self.sheetname = sheetname
        self.savefile = savefile
        self.dataframe = dataframe
        # 최대열 미지정시 서식파일의 마지막 열
        self.max_row = self._get_maxrow() if max_row is None else max_row
        self._mappings = mappings
        self._border_settings = border_settings
        self._sheetname = sheetname
        self._repeats = len(self.dataframe)

    @property
    def set_mapping(self, mappings: list) -> None:
        """
        mapping 설정 수정
        param mappings: list(dict), 데이터 필드 및 좌표 지정, 처리함수의 dictionary
        """
        self._mappings = mappings

    @property
    def set_bordersettings(self, border_settings:list) -> None:
        """
        border 설정 수정
        param border_settings: list(dict), 서식의 경계선 정의
        """
        self._border_settings = border_settings

    @classmethod
    def _get_maxrow(self) -> int:
        """
        서식 파일의 마지막 열
        """
        template = openpyxl.load_workbook(self.template)
        sheet = template.worksheets[0] if self._sheetname is None else template[self._sheetname]
        template.close()
        return sheet.max_row

    def data_mapping(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """
        데이터 매핑 정의에 따른 데이터 작성
        param sheet: openpyxl.worksheet.worksheet.Worksheet, 데이터가 작성될 시트 객체
        """

        for datamap in self._mappings:

            ranges = datamap.get('address')
            fields = datamap.get('fields')
            callback = datamap.get('callback')
            kargs = datamap.get('kargs')

            rng = ranges

            for i in range(self._repeats):             
                if callback is None:
                    # 단순 데이터 기록
                    oa.set_data(sheet, rng, self.dataframe.loc[i, fields])
                else:
                    # 특정 함수를 통과 후 기록(필드값 합치기, 이미지 입력.. )
                    # shp2report_callbacs.py에 정의
                    if kargs is None:
                        callback(sheet=sheet, rng=rng, index=i, dataframe=self.dataframe, fields=fields)
                    else:
                        callback(sheet=sheet, rng=rng, index=i, dataframe=self.dataframe, fields=fields, **kargs)
                rng = oa.cell_shift(rng, row_shift=self.max_row)


    def report(self) -> None:
        """
        실제 레포트 작성
        """
        # mappings가 정의되지 않은 경우 에러발생
        if self._mappings is None:
            raise Exception("Mapping Not Defined")
        
        # 데이터 수에 맞도록 서식 복사
        wb, ws = oa.copy_templates2(self.template, sht_name=self._sheetname, target_path=self.savefile, 
                                    repeats=self._repeats, max_row=self.max_row, border_settings=self._border_settings)
        # 데이터 작성
        self.data_mapping(sheet=ws)

        # 저장
        wb.save(self.savefile)
        wb.close()


if __name__ == "__main__":

    from shp2report_callbacks import str_add, str_deco, insert_image

    border_setting =[{"rng": "A3:D4","edges": ["all"], "border_style": "thin", "reset": True },  # 전체 가는 실선
                    {"rng": "A3:A5","edges": ["all"], "border_style": "thin",  "reset": False },
                    {"rng": "A6:D11","edges": ["all"], "border_style": "thin",  "reset": False },
                    {"rng": "A11:D11","edges": ["outer"], "border_style": "medium", "reset": False },  # 외곽선, 중간 실선
                    {"rng": "A3:D11","edges": ["outer"], "border_style": "medium", "reset": False }
                     ]
    mapping = [ {'fields': 'Name', 'address': 'A1', 'callback': str_deco, 'kargs':{'prefix': 'Temporary data of (', 'postfix': ')'}}, ## f'Temporaty data of ({data})'
            {'fields': 'ID', 'address': 'D2'},
            {'fields': 'Name', 'address': 'B3'},
            {'fields': 'Age', 'address': 'D3'},
            {'fields': 'Address', 'address': 'B4'},
            {'fields':['Bigo', 'Name', 'Age'], 'address': 'B11', 'callback': str_add, 'kargs':{'delim': ' '}},   ## f'{Bigo} {Name} {Age}'
            {'fields': 'Map', 'address': "A6:D10", 'callback': insert_image, 'kargs':{'keep_ratio': True}}]  ## insert image
    
    df = pd.read_excel('data.xlsx', dtype={'ID':str})
    df.loc[:, 'Map'] = 'abcd.jpg'

    my_report = ReportFromDataframe(template='templates.xlsx', savefile = 'data_report_example.xlsx', dataframe=df, max_row=12, border_settings=border_setting, mappings=mapping)
    my_report.report()
