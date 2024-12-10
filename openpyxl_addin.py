import openpyxl
from openpyxl.styles import Border, Side, Alignment
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.drawing.image import Image
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU, EMU_to_pixels, EMU_to_cm, pixels_to_points, inch_to_EMU
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.cell import get_column_letter, column_index_from_string, coordinate_from_string, range_boundaries
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from copy import copy
import ctypes
import shutil
import numpy as np
from datetime import datetime, timedelta

from tkinter import Tk
from tkinter.font import Font as TkFont
import os


MIN_WIDTH = 2

def _get_text_size(text: str, font_family: str = 'Arial', font_size: int = 10, bold: bool = False, italic: bool = False) -> int:
    """Get the screen width of a text based on Font Type, Font Size and Font Weight

    Args:
        text (str): Text for which to calculate the screen width
        font_family (str, optional): Font family. Defaults to 'Arial'.
        font_size (int, optional): Font size. Defaults to 10.
        bold (bool, optional): If bold or not. Defaults to False.

    Returns:
        int: Screen width of the text
    """
    root = Tk()  # Needed to estimate the width.
    reference_font = TkFont(family='Calibri', size=11)
    font_weight = 'bold' if bold else 'normal'
    font_italic = 'italic' if italic else 'roman'
    font_var = TkFont(family=font_family, size=font_size, weight=font_weight, slant=font_italic)
    width = font_var.measure(text) / reference_font.measure('0')
    root.destroy()  # Destroy the created window
    return width

def set_border(rng: tuple, edges: list=['all'], border_style: str =None, color: str ='000000', reset: bool =True) -> None:
    """Description: 범위내 테두리선의 색상 및 두께를 지정
       param rng: 적용할 범위, ex) ws['A2:D8']
       param edges: list, 적용할 테두리 선택, [all, top, bottom, left, right, horizontal, vertical, inner_horizontal, inner_vertical, outer]
       border_style: 테두리선 종류, [None, 'dashDot','dashDotDot', 'dashed','dotted',
                                    'double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot',
                                    'mediumDashed', 'slantDashDot', 'thick', 'thin']
       color: 적용할 색상, hex code
       reset: 기존 스타일 사용 여부
    """
    b_style = None if border_style is None else Side(border_style=border_style, color=color)

    for i, rows in enumerate(rng):
        for j, cell in enumerate(rows):
            border = Border() if reset else Border(left=cell.border.left,
                                                    right=cell.border.right,
                                                    top=cell.border.top,
                                                    bottom=cell.border.bottom)
            for edge in edges:
                if i == 0 :
                    if edge in ['all', 'top', 'horizonal', 'outer']:
                        border.top = b_style
                    if edge in ['all', 'horizontal', 'inner_horizontal']:
                        border.bottom = b_style
                if 0 < i < len(rng)-1:
                    if edge in ['all', 'horizontal', 'inner_horizontal']:
                        border.bottom = b_style
                if i == len(rng)-1:
                    if edge in ['all', 'bottom', 'horizontal', 'outer']:
                        border.bottom = b_style
                    if edge in ['all', 'horizontal', 'inner_horizontal']:
                        border.top = b_style
                if j == 0:
                    if edge in ['all', 'left', 'vertical', 'outer']:
                        border.left = b_style
                    if edge in ['all', 'vertical', 'inner_vertical']:
                        border.right = b_style
                if 0 < j < len(rows):
                    if edge in ['all', 'vertical', 'inner_vertical']:
                        border.left = b_style
                if j == len(rows)-1:
                    if edge in ['all', 'right', 'vertical', 'outer']:
                        border.right = b_style
                    if edge in ['all', 'vertical', 'inner_vertical']:
                        border.left = b_style

            cell.border = border

def set_font(rng: tuple, **kargs) -> None:
    """ 
    범위 내 폰트 적용 
    param rng: (tuple) ex. ws['A1:K1']
    param kargs
          b: bool, 굵게
          i: bool, 이탤릭체
          sz: float, 크기
          name: str, 폰트명
          color: openpyxl.styles.colors.Color, 색상(ex. 'FF00DD')
          strike: bool, 취소선
          u: one of {'singleAccounting', 'doubleAccounting', 'double', 'single'}, 실선, 이중실선, 실선(회계용), 이중실선(회계용)
          vertAlign: one of {'baseline', 'subscript', 'superscript'}, 첨자
    """
    for rows in rng:
        for cell in rows:
            f = copy(cell.font)
            f.b = kargs.get('b')
            f.i = kargs.get('i')
            f.sz = kargs.get('sz')
            f.name = kargs.get('name')
            f.color = kargs.get('color')
            f.strike = kargs.get('strike')
            f.u = kargs.get('u')
            f.vertAlign = kargs.get('vertAlign')

            cell.font = f

def set_alignment(rng: tuple, **kargs) -> None:
    """ 
    정렬 설정 
    param kargs
          horizontal: one of {'fill', 'general', 'justify', 'center', 'left', 'centerContinuous', 'distributed', 'right'}  
          indent: float         
          shrink_to_fit: bool, 셀맞춤
          textRotation: integer (0<= value <=180)
          vertical: one of {'justify', 'center', 'distributed', 'top', 'bottom'}
          wrapText: bool, 자동줄바꿈
    """
    for rows in rng:
        for cell in rows:
            cell.alignment = Alignment(**kargs)

def copy_templates(template_sheet: openpyxl.worksheet.worksheet.Worksheet, target_sheet: openpyxl.worksheet.worksheet.Worksheet,
               repeats: int=4, max_row: int=None, border_settings: list = None) -> None:
    """
    지정 수량 만큼 서식 시트를 타겟 시트에 복사
    params   template_sheet: openpyxl.worksheet.worksheet.Worksheet, 서식 시트
             target_sheet: openpyxl.worksheet.worksheet.Worksheet, 타겟 시트
             repeats: int, 복사 수(페이지 수)
             max_row: int, 서식 시트의 최대 행수(서식 시트의 범위가 A1:D15라면 max_row=15)
             border_settings: list[dict], set_border의 파라미터를 가지는 dict의 리스트, 단, rng는 셀범위만 입력
    """
    
    max_row = template_sheet.max_row if max_row is None else max_row
    mergecells = template_sheet.merged_cells

    _copy_cells(template_sheet, target_sheet, repeats, max_row)  # copy all the cel values and styles
    _copy_sheet_attributes(template_sheet, target_sheet, repeats, max_row)

    ## 병합된 셀 반복 적용
    for cell in mergecells:        
        cell_add = CellRange(cell.coord)
        for _ in range(1, repeats):
            cell_add.shift(row_shift=max_row)
            target_sheet.merge_cells(cell_add.coord)

    ## 테두리 설정 반복 적용
    if not border_settings is None:
        for border_setting in border_settings:
            border_rng = CellRange(border_setting['rng'])
            for i in range(repeats):
                if i > 0:                     
                    border_rng.shift(row_shift=max_row)
                border_setting['rng'] = target_sheet[border_rng.coord]
                set_border(**border_setting)

def copy_templates2(template: str, sht_name: str=None, target_path: str=None, 
                    repeats: int=4, max_row: int=None, border_settings: list = None) -> tuple:
    """
    지정 수량 만큼 서식 시트를 타겟 시트에 복사, 파일복사 후 복사된 파일에서 수량만큼 서식복사
    params   template: str, 서식(.xlsx)
             target_path: str, 타겟파일(.xlsx)
             repeats: int, 복사 수(페이지 수)
             max_row: int, 서식 시트의 최대 행수(서식 시트의 범위가 A1:D15라면 max_row=15)
             border_settings: list[dict], set_border의 파라미터를 가지는 dict의 리스트, 단, rng는 셀범위만 입력
    return  workbook, worksheet
    """
    # 서식파일을 다른 이름으로 저장
    target_path = template.replace('.xlsx', '_aldfjladf.xlsx') if target_path is None else target_path
    shutil.copy(template, target_path)
    wb = openpyxl.load_workbook(target_path)
    if sht_name is None:
        ws = wb.worksheets[0]
    else:
        ws = wb[sht_name]

    max_row = ws.max_row if max_row is None else max_row
    
    mergecells = copy(ws.merged_cells)

    _copy_templates2(ws, repeats, max_row)  # copy all the cel values and styles

    ## 병합된 셀 반복 적용
    for cell in mergecells:        
        cell_add = CellRange(cell.coord)
        for _ in range(1, repeats):
            cell_add.shift(row_shift=max_row)
            ws.merge_cells(cell_add.coord)

    ## 테두리 설정 반복 적용
    if not border_settings is None:
        for border_setting in border_settings:
            border_rng = CellRange(border_setting['rng'])
            for i in range(repeats):
                if i > 0:                     
                    border_rng.shift(row_shift=max_row)
                border_setting['rng'] = ws[border_rng.coord]
                set_border(**border_setting)

    return wb, ws

def _copy_sheet_attributes(template_sheet: openpyxl.worksheet.worksheet.Worksheet, target_sheet: openpyxl.worksheet.worksheet.Worksheet, repeats: int=1, max_row: int=None) -> None:
    """ 시트의 속성 복사 """
    target_sheet.sheet_format = copy(template_sheet.sheet_format)
    target_sheet.sheet_properties = copy(template_sheet.sheet_properties)
    target_sheet.merged_cells = copy(template_sheet.merged_cells)
    target_sheet.page_margins = copy(template_sheet.page_margins)
    target_sheet.freeze_panes = copy(template_sheet.freeze_panes)

    max_row = template_sheet.max_row if max_row is None else max_row

    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). 
    # So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(template_sheet.row_dimensions)):
        for nth in range(repeats):
            target_sheet.row_dimensions[1 + rn + nth*max_row] = copy(template_sheet.row_dimensions[1 + rn])

    if template_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(template_sheet.sheet_format.defaultColWidth)

    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, _ in template_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(template_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(template_sheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not only the hidden property
        target_sheet.column_dimensions[key].width = copy(template_sheet.column_dimensions[key].width) # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(template_sheet.column_dimensions[key].hidden)

def _copy_templates2(sheet: openpyxl.worksheet.worksheet.Worksheet, repeats: int=1, max_row: int=None) -> None:
    for i in range (1, max_row + 1):
        for j in range (1, sheet.max_column + 1):
            # Assign source cell value
            cell_src = sheet.cell(row = i, column = j)
            r_dim = sheet.row_dimensions[i]
            #print(cell_src.fill)
            # paste values and styles to the destination sheet
            for n in range(1, repeats):
                sheet.cell(row = i + n*max_row, column = j).value = cell_src.value
                sheet.cell(row = i + n*max_row, column = j).fill = copy(cell_src.fill)
                sheet.cell(row = i + n*max_row, column = j).font  = copy(cell_src.font )
                sheet.cell(row = i + n*max_row, column = j).border  = copy(cell_src.border )
                sheet.cell(row = i + n*max_row, column = j).alignment  = copy(cell_src.alignment )
                sheet.row_dimensions[i+n*max_row] = r_dim

def _copy_cells(template_sheet: openpyxl.worksheet.worksheet.Worksheet, target_sheet: openpyxl.worksheet.worksheet.Worksheet, repeats: int=1, max_row: int=None) -> None:
    """ 셀의 값과 속성을 복사 """
    for (row, col), source_cell in template_sheet._cells.items():
        for nth in range(repeats):
            target_cell = target_sheet.cell(column=col, row=row + nth*max_row)
            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type

            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

def _ch_px(v: float) -> float:
    """
    Convert between Excel character width and pixel width.
    """
    return v * 8

def _get_windows_dpi() -> int:
    hdc = ctypes.windll.user32.GetDC(0)
    dpi = ctypes.windll.gdi32.GetDeviceCaps(hdc, 88)  # 88 corresponds to LOGPIXELSX (horizontal DPI)
    ctypes.windll.user32.ReleaseDC(0, hdc)
    return dpi

def copyRange(sheet, rngstr: str) -> dict:
    """
    지정된 시트의 범위의 값을 복사, 병합된 셀정보 포함
    param sheet: openpyxl.worksheet.worksheet.Worksheet, 시트
          rngstr: str, 셀 범위. ex: 'B2:F6'
    return dict{'data': 셀 값, 'basecoord': 범위의 좌상단 좌표(col, row), 'in_merge': 범위내 병합셀}
    """
    rng = CellRange(rngstr)
    c_left, r_top, c_right, r_bottom = range_boundaries(rngstr)
    rangeSelected = []
    # Loops through selected Rows
    for i in range(r_top, r_bottom + 1, 1):
        # Appends the row to a RowSelected list
        rowSelected = []
        for j in range(c_left, c_right + 1, 1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        # Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
    in_merged_cells = []
    merged_cells = copy(sheet.merged_cells)
    # 병합 셀 정보
    for m_cell in merged_cells:
        if rng.issuperset(m_cell):
            in_merged_cells.append(m_cell)
    return {'data':rangeSelected, 'basecoord':(c_left, r_top), 'in_merge': in_merged_cells}

def pasteRange(sheet, targetcell: str, copied: dict) -> None:
    """
    지정된 시트에 복사한 값 삽입, 병합된 셀정보 포함
    param sheet: openpyxl.worksheet.worksheet.Worksheet, 붙여넣을 시트
          targetcell: str, 붙여넣을 셀(하나의 셀). ex: 'F6'
          copied: dict, copyRange에 의해 복사된 정보
    return None
    """
    copied_data = copied['data']
    r, c = np.shape(copied_data)
    startCol = column_index_from_string(coordinate_from_string(targetcell)[0])
    startRow = coordinate_from_string(targetcell)[1]

    c_shift, r_shift = np.array((startCol, startRow)) - np.array(copied['basecoord'])

    endCol = startCol + c - 1
    endRow = startRow + r - 1
    countRow = 0
    for i in range(startRow, endRow + 1, 1):
        countCol = 0
        for j in range(startCol, endCol + 1, 1):
            sheet.cell(row=i, column=j).value = copied_data[countRow][countCol]
            countCol += 1
        countRow += 1
    # 셀 병합
    for c in copied['in_merge']:
        sheet.merge_cells(cell_shift(c.coord, row_shift = r_shift, col_shift= c_shift))

# 열 너비 자동 맞춤
def autofit_column_size(worksheet: openpyxl.worksheet.worksheet.Worksheet, columns: str=None) -> None:
    """
    열 너비 자동 맞춤
    param worksheet: openpyxl.worksheet.worksheet.Worksheet, 해당 시트
          columns: 대상 열(ex. 'A:D, F, H:J'), default=None(for All columns)
          return None
    """
    dims = dict()

    # 열에 대한 문자열을 컬럼 문자열로 변환
    column_indexes = [get_column_letter(row[0].column) for row in worksheet.columns]

    for c in column_indexes:
        dims[c] = worksheet.column_dimensions[c].width

    if not columns is None:
        column_indexes = parse_column_representation(columns)

    # 해당 열의 각 셀에 대해 폰트 크기, 종류 등을 반영한 최대 너비를 dims에 저장
    for index in column_indexes:
        value = _text_width(worksheet, index)
        dims[index] = value

    # 각 열에 대해 너비 반영    
    for key in dims.keys():
        worksheet.column_dimensions[key].width = dims.get(key)

def parse_column_representation(col_repr: str) -> list:
    """
    컬럼 표현식을 개별 컬럼문자열로 반환
    param col_repr: str, 컬럼 표현식(ex. 'A, C:E, G:I')
    return list, 개별 컬럼문자 리스트(ex. ['A', 'C', 'D', 'E', 'G', 'H', 'I'])
    """
    if (not col_repr) or (not isinstance(col_repr, str)):
        return None

    columns = col_repr.strip().split(',') if ',' in col_repr else [col_repr.strip()]

    # 콜론(:)으로 지정된 열범위를 개별 문자로 생성
    _indexes = []
    for col in columns:
        if ':' in col:
            _start, _end = col.split(':')
            _c1, _c2 = column_index_from_string(_start.strip()), column_index_from_string(_end.strip())
            _min, _max = min((_c1, _c2)), max((_c1, _c2))
            for i in range(_min, _max+1):
                _indexes.append(get_column_letter(i))
        else:
            _indexes.append(col.strip())
    return _indexes

def insert_image(sheet: openpyxl.worksheet.worksheet.Worksheet, rng: str, src: str, keep_ratio: bool = True, margin_px: int = 2, **kargs) -> None:
    """
        지정된 범위에 이미지 삽입
        params sheet: openpyxl.worksheet.worksheet.Worksheet, 삽입하고자 하는 시트
               rng: str, 범위, 예) "A2:C3"
               src: str, 삽입할 이미지 경로
               keep_ratio: 원본 이미지 비율 유지 여부
               margin_px: 여백 지정
        return None
    """

    # src가 존재하지 않거나 파일이 아니면 처리없이 리턴
    if not (os.path.exists(src) and os.path.isfile(src)):
        return

    # DimensionHolder로 열너비 재설정
    column_width = [17.625, 8.625, 3.125, 3.125, 2.505, 2.505, 2.505, 2.875, 2.505, 1.625, 2.005, 1.755, 1.755, 1.755, 1.755, 1.755, 1.755, 1.755, 1.755, 1.755, 1.755, 1.755, 1.755, 1.755, 1.755, 1.755, 1.755, 1.755, 1.755, 1.755, 1.755, 1.755]
    dim_holder = DimensionHolder(worksheet=sheet)
    for i, w in enumerate(column_width):
        dim_holder[get_column_letter(i+1)] = ColumnDimension(worksheet=sheet, min=i+1, max=i+1, width=w)
    sheet.column_dimensions = dim_holder
    
    # image_path to Image object
    image = Image(src)

    # cell index starts from 0, row index starts from 1
    c_left, r_top, c_right, r_bottom = range_boundaries(rng)
   
    if keep_ratio:  ## 이미지 비율 유지
        # cell width to accumulated list(unit: EMU)
        r_to_EMU = lambda x: pixels_to_EMU(_ch_px(x)) - 4 / (_get_windows_dpi() / 72)
        widths = [r_to_EMU(sheet.column_dimensions[get_column_letter(column+1)].width) for column in range(c_left, c_right)]
        accum_width = [sum(widths[:i+1]) for i in range(len(widths))]
        cw = accum_width[-1]

        # cell height to accumulated list(unit: EMU)
        heights = [(16.5 if sheet.row_dimensions[row].height is None else sheet.row_dimensions[row].height) for row in range(r_top, r_bottom+1)]
        h_to_EMU = lambda x: pixels_to_EMU(x * _get_windows_dpi() / 72)
        heights = list(map(h_to_EMU, heights))
        accum_height = [sum(heights[:i+1]) for i in range(len(heights))]
        ch = accum_height[-1]

        # original image width, height to EMU
        iw, ih = pixels_to_EMU(image.width), pixels_to_EMU(image.height)

        # calculate maximum scale ratio to fit in range
        ratio = cw/iw if iw/cw > ih/ch else ch/ih
        iw, ih = pixels_to_EMU(image.width) * ratio - pixels_to_EMU(margin_px), pixels_to_EMU(image.height) * ratio - pixels_to_EMU(margin_px)  ## EMU

        # calculate margins(offset) 
        roff = (ch-ih)/2 
        coff = (cw-iw)/2 

        # find anchor bias and recalculate offset
        for i, val in enumerate(accum_width):
            c_bias = i
            
            if val > coff:
                if len(accum_width) == 1:
                    break

                if i==0:
                    if coff > val/2:                        
                        c_bias += 1
                        coff -= val                    
                else:
                    if coff > (val + accum_width[i-1])/2:
                        c_bias += 1
                        coff -= val
                    else:
                        coff -= accum_width[i-1]
                break
        
        for i, val in enumerate(accum_height):
            r_bias = i
            if val > roff:
                if i > 0:
                    roff -= accum_height[i-1]
                break

        # set anchor
        marker = AnchorMarker(col=c_left-1 + c_bias, colOff=coff, row=r_top-1 + r_bias, rowOff=roff)  ## offset <- EMU
        size = XDRPositiveSize2D(iw, ih)
        anchor = OneCellAnchor(_from=marker, ext=size)

    else:  ## 이미지 비율을 셀크기에 맞게
        offset = pixels_to_EMU(margin_px)  # margin
        _from = AnchorMarker(
            col=c_left - 1,
            row=r_top - 1,
            colOff=offset,
            rowOff=offset,
        )
        to = AnchorMarker(
            col=c_right,
            row=r_bottom,
            colOff=-offset,
            rowOff=-offset,
        )
        anchor = TwoCellAnchor(_from=_from, to=to)

    image.anchor = anchor

    # Add the image to the worksheet
    sheet.add_image(image)


def set_data(sheet: openpyxl.worksheet.worksheet.Worksheet, rng:str, data: any) -> None:
    sheet[rng] = data

def cell_shift(rng: str, **kargs) -> str:
    """
    셀 주소를 지정된 값만큼 이동된 좌표 반환
    param rng: str, 셀 주소, ex) 'A2', 'A6:C9'
          kargs: 
            row_shift: row 이동량, col_shift: column 이동량
    return str, 이동된 주소값, ex) 'F12', 'D3:F6'
    """
    _ranges = CellRange(rng)
    rowshift = 0 if kargs.get('row_shift') is None else kargs.get('row_shift')
    colshift = 0 if kargs.get('col_shift') is None else kargs.get('col_shift')
    _ranges.shift(row_shift=rowshift, col_shift=colshift)
    return _ranges.coord

def _text_width(sht: openpyxl.worksheet.worksheet.Worksheet, col_ref: str) -> float:
    """
    column auto-fit을 위한 함수, 셀 값과 폰트를 고려하여 최적의 컬럼 너비를 반환
    param sht: 입력 시트
          col_ref: str, 컬럼 문자, 'A'
    return float
    """
    rng= sht[col_ref]

    text_width = []
    for c in rng:
        if c.coordinate in sht.merged_cells:
            continue

        if not c.value:
            continue

        fs = copy(c.font)
        name, size, bold, italic = fs.name, int(fs.size), fs.bold, fs.i
        text = str(c.value)
        text_width.append(_get_text_size(text, name, size, bold, italic))
    # print(text_width)
    # 열너비는 선형회귀를 돌려서 새로 지정
    return max(text_width) * 0.8837 + 1.8 if text_width else MIN_WIDTH

def copy_row_with_merge(ws: openpyxl.worksheet.worksheet.Worksheet, source_row: int, target_row: int, repeat: int):
    """
    특정 행을 복사하여 지정된 행에 반복 삽입하는 함수
    
    :param ws: 작업 중인 워크시트
    :param source_row: 복사할 원본 행 번호
    :param target_row: 삽입할 대상 행 번호
    :param repeat: 반복 횟수
    """
    # 원본 행의 병합된 셀 찾기
    merged_cells_in_source = [
        merged_range for merged_range in ws.merged_cells.ranges 
        if merged_range.min_row == source_row and merged_range.max_row == source_row
    ]
    # 원본 행의 데이터와 서식 복사
    row_data = []
    row_styles = []

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=source_row, column=col)
        row_data.append(cell.value)
        row_styles.append({
            'font': copy(cell.font),
            'fill': copy(cell.fill),
            'border': copy(cell.border),
            'alignment': copy(cell.alignment),
            'number_format': cell.number_format
        })

    # 지정된 횟수만큼 반복
    for _ in range(repeat):
        # 대상 행에 새 행 삽입
        ws.insert_rows(target_row)
        # 복사한 데이터와 스타일 적용
        for col in range(1, len(row_data)+1):
            new_cell = ws.cell(row=target_row, column=col)
            new_cell.value = row_data[col-1]
            
            # 스타일 복사
            new_cell.font = row_styles[col-1]['font']
            new_cell.fill = row_styles[col-1]['fill']
            new_cell.border = row_styles[col-1]['border']
            new_cell.alignment = row_styles[col-1]['alignment']
            new_cell.number_format = row_styles[col-1]['number_format']


    # 병합된 셀 복사
    for merged_range in merged_cells_in_source:
        for n in range(1, repeat+1): 
            # 병합 범위의 경계 얻기
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            
            # 새 행 계산
            new_min_row = min_row + n
            new_max_row = max_row + n
            
            # 새 병합 범위 생성
            new_merge_range = f"{get_column_letter(min_col)}{new_min_row}:{get_column_letter(max_col)}{new_max_row}"
            
            # 병합
            ws.merge_cells(new_merge_range)

def format_date_to_korean(datetime_str):
    # 문자열에서 날짜 부분 추출
    date_str = datetime_str.split(" ")[0]
    
    # 날짜를 YYYY, MM, DD로 분리
    year, month, day = date_str.split("-")
    
    # 한국어 형식으로 변환
    formatted_date = f"{year}년 {int(month):02d}월 {int(day):02d}일"
    return formatted_date

def convert_angle_to_decimal(angle_str, n=4) -> tuple:
    # 입력 문자열에서 각도, 분, 초를 분리, 초는 n자리에서 오사오입
    if not angle_str.endswith('"'):
        angle_str += '"'
    angle_str = angle_str.replace("˚", "˚ ").replace("'", "' ")
    angle_str = angle_str.replace("  ", " ")
    
    degrees, minutes, seconds = angle_str.split(' ')
    
    # 각도와 분은 정수로 변환
    degrees = int(degrees[:-1])  # 마지막 글자 '˚' 제거
    minutes = int(minutes[:-1])  # 마지막 글자 ''' 제거
    seconds = float(seconds[:-1])  # 마지막 글자 '"' 제거

    # 초를 소수점 4째 자리에서 반올림
    seconds_rounded = np.round(seconds, n)
    
    # 변환된 값 반환
    return degrees, minutes, seconds_rounded

def convert_decimal_to_angle(degrees, minutes, seconds) -> str:
    # 각도, 분, 초를 문자열 형식으로 변환
    angle_str = f"{degrees}˚ {minutes}' {seconds:.4f}\""
    return angle_str
    
def convert_decimal_to_roundup_angle(angle_str, n=4) -> str:    
    # 각도, 분, 초를 문자열 형식으로 변환
    degrees, minutes, seconds = convert_angle_to_decimal(angle_str, n)    
    angle_str = f"{degrees}˚ {minutes}' {seconds:.4f}\""
    return angle_str


if __name__ == '__main__':

    ##############################################
    # wb = Workbook()
    # ws = wb.active
    
    ### 폰트 설정
    # font_args = {'sz': 16, 'color': 'ff0000', 'u': 'doubleAccounting', 'scheme':'minor'}
    # set_font(ws["B2:E6"], sz=16, color ='00ff00', u='doubleAccounting', scheme='minor')
    # set_font(ws["B2:E6"], **font_args)

    ### 테두리 설정
    # set_border(ws["B2:E6"], ['inner_horizontal', 'inner_vertical'], "hair", '0000ff', True)
    # set_border(ws["B2:E6"], ['outer'], "medium", '000000', False)

    # wb.save('abc.xlsx')

    ###############################################
    # xl_report = "report.xlsx"
    # template = "data.xlsx"

    # # df = pd.read_excel('data.xlsx', sheet_name='Sheet1')

    # new_workbook = openpyxl.Workbook()
    # new_sheet = new_workbook.active

    # # template_wb = openpyxl.load_workbook(template)
    # # template_ws = template_wb.worksheets[0]

    # border_setting =[{"rng": "A3:D4","edges": ["all"], "border_style": "thin", "reset": True },
    #                  {"rng": "A3:A5","edges": ["all"], "border_style": "thin",  "reset": False },
    #                  {"rng": "A6:D11","edges": ["all"], "border_style": "thin",  "reset": False },
    #                  {"rng": "A11:D11","edges": ["outer"], "border_style": "medium", "reset": False },
    #                  {"rng": "A3:D11","edges": ["outer"], "border_style": "medium", "reset": False }
    #                  ]
    # # copy_templates(template_ws, new_sheet, repeats=5, max_row=12, border_settings = border_setting)
    
    # savefile = 'report_aldladsjf.xlsx'
    # wb = copy_templates2(template, target_path=savefile, repeats=5, max_row=12, border_settings = border_setting)
    
    # rng = CellRange("A6:D10")
    # for i in range(5):      
    #     insert_image(wb[0], rng.coord, 'abcd.jpg', keep_ratio=True)
    #     rng.shift(row_shift=12)

    # ws = wb.active
        
    # wb.save(savefile)  
    # wb.close() 

    #############################################   
    template = "data.xlsx"

    wb = openpyxl.load_workbook(template)
    ws =  wb.worksheets[0]


    # copied = copyRange(ws, "A2:d8")
    # pasteRange(ws, "F2", copied)

    # copied = copyRange(ws, "g2:h5")
    # pasteRange(ws, "g16", copied)

    set_font(ws['A1:K1'], name='HY헤드라인M', b=True, sz=10, i=True)

    wb.save('data1.xlsx')
    wb.close()
