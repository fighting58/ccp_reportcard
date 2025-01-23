import sys
import os
import geopandas as gpd
import pandas as pd
import numpy as np
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QFileDialog, QPushButton, QLineEdit,
                                QRadioButton, QTableWidget, QToolBar, QButtonGroup, QStyledItemDelegate,
                                QHeaderView, QTableWidgetItem, QStatusBar, QLabel, QFrame, QScrollArea,
                                QCheckBox, QVBoxLayout, QHBoxLayout, QSpacerItem, QDockWidget, QGroupBox, 
                                QSizePolicy, QAbstractItemView, QMenu, QLabel, QComboBox, QInputDialog, 
                                QDialog, QGraphicsDropShadowEffect, QTextBrowser)
from PySide6.QtCore import Qt, QRect, Signal, QTimer, Slot, QSize, QFile, QIODevice, QTextStream
from PySide6.QtGui import QFontMetrics, QKeySequence, QPainter, QPen, QColor, QIcon, QAction, QFont
import resources
import pickle   
from geometric_search import find_attributes_containing_point
from shp2report import ReportFromDataframe
from shp2report_callbacks import insert_image, str_add, str_deco, hangul_date, toBL, osa
from cif_converter import CifGeoDataFrame
import pickle
from CodeDownload_codegokr import CodeGoKr
from custom_image_editor import ImageEditor
from rename_image_with_tr import DialogRenameImage
from openpyxl import load_workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl_addin import set_alignment, set_border, set_font, copy_row_with_merge, format_date_to_korean, convert_decimal_to_roundup_angle
from datetime import datetime, timedelta
import random
from QCustomModals import QCustomModals
from settings import Settings
from environment_manage import EnvironmentManager
from ui_splash_screen import Ui_SplashScreen
import webbrowser
from vworldmap_dialog import VWorldMapViewer


# global value
splash_timer_value = 0

class CustomToggleButton(QWidget):
    stateChanged = Signal(bool)  # 상태 변경 시그널

    def __init__(self, parent=None):
        super().__init__(parent)
        self.initUI()

    def initUI(self):
        layout = QHBoxLayout()
        self.setLayout(layout)

        # 토글 버튼 생성
        self.toggleButton = QPushButton(self)
        self.toggleButton.setCheckable(True)
        self.toggleButton.setFixedSize(120, 40)  # 버튼 크기 조정
        self.toggleButton.toggled.connect(self.onToggle)
        self.toggleButton.setChecked(True)
        self.toggleButton.setStyleSheet("border: none")

        # 레이아웃에 위젯 추가
        layout.addWidget(self.toggleButton)
        layout.setAlignment(Qt.AlignLeft)
        layout.setContentsMargins(0, 0, 0, 0)

    def onToggle(self, checked):
        if checked:
            self.toggleButton.setIcon(QIcon(':resources/icons/toggle-off-circle.svg'))
            self.toggleButton.setText("사진 편집 모드")

        else:
            self.toggleButton.setIcon(QIcon(':resources/icons/toggle-on-circle.svg'))
            self.toggleButton.setText("표 편집 모드")
        self.stateChanged.emit(checked)

    def isChecked(self):
        return self.toggleButton.isChecked()

    def setChecked(self, checked):
        self.toggleButton.setChecked(checked)


class CustomTableWidget(QTableWidget):
    item_double_clicked = Signal(int, str, str, str)

    def __init__(self, parent=None):
        super().__init__()        
        self.setMouseTracking(True)
        self.start_item = None
        self.end_item = None
        self.dragging = False
        self.fill_handle_size = 6
        self.drag_rect = None
        self._is_edit_mode = True
        self.setup_headers()
        self.itemDoubleClicked.connect(self.on_item_double_clicked)
        
    @property
    def mode(self):
        return self._is_edit_mode
    @mode.setter
    def set_mode(self, mode: bool):
        self._is_edit_mode = mode

    def setup_headers(self):
        horizontal_header = self.horizontalHeader()
        vertical_header = self.verticalHeader()
        
        horizontal_header.setContextMenuPolicy(Qt.CustomContextMenu)
        vertical_header.setContextMenuPolicy(Qt.CustomContextMenu)
        
        horizontal_header.customContextMenuRequested.connect(self.show_column_menu)
        vertical_header.customContextMenuRequested.connect(self.show_row_menu)

    def on_item_double_clicked(self, item):
        if not self.mode:
            row = item.row()
            num = self.item(row, 0).text()  # 점번호
            path = self.item(row, 18).text()  # 사진파일(경로)
            name = self.item(row, 19).text()  # 사진파일명

            # 신호 발송
            self.item_double_clicked.emit(row, num, path, name)

    def mousePressEvent(self, event):
        item = self.itemAt(event.position().toPoint())
        if item and self.is_on_fill_handle(event.position().toPoint(), item):
            self.start_item = item
            self.dragging = True
        else:
            super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self.dragging:
            self.end_item = self.itemAt(event.position().toPoint())
            if self.end_item:
                self.update_drag_rect()
            self.viewport().update()
        else:
            super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if self.dragging:
            self.dragging = False
            if self.start_item and self.end_item:
                self.fill_items()
            self.drag_rect = None
            self.viewport().update()
        else:
            super().mouseReleaseEvent(event)

    def show_column_menu(self, pos):
        column = self.horizontalHeader().logicalIndexAt(pos)
        menu = QMenu()
        delete_action = menu.addAction("Delete Column")
        action = menu.exec(self.horizontalHeader().mapToGlobal(pos))
        
        if action == delete_action:
            self.removeColumn(column)

    def show_row_menu(self, pos):
        row = self.verticalHeader().logicalIndexAt(pos)
        menu = QMenu()
        delete_action = menu.addAction("Delete Row")
        action = menu.exec(self.verticalHeader().mapToGlobal(pos))
        
        if action == delete_action:
            self.removeRow(row)

    def fill_items(self):
        start_row = self.start_item.row()
        start_col = self.start_item.column()
        end_row = self.end_item.row()
        end_col = self.end_item.column()
        value = self.start_item.text()

        row_step = 1 if start_row < end_row else -1
        col_step = 1 if start_col < end_col else -1        

        for row in range(start_row, end_row + 1 * row_step, row_step):
            for col in range(start_col, end_col + 1 * col_step, col_step):
                if row != start_row or col != start_col:
                    item = self.item(row, col)
                    if not item:
                        item = QTableWidgetItem()
                        item.setTextAlignment(Qt.AlignCenter)
                        self.setItem(row, col, item)
                    item.setText(value)

    def is_on_fill_handle(self, pos, item):
        rect = self.visualItemRect(item)
        handle_rect = QRect(rect.right() - self.fill_handle_size, rect.bottom() - self.fill_handle_size, self.fill_handle_size, self.fill_handle_size)
        return handle_rect.contains(pos)

    def update_drag_rect(self):
        start_rect = self.visualItemRect(self.start_item)
        end_rect = self.visualItemRect(self.end_item)
        self.drag_rect = start_rect.united(end_rect)

    def paintEvent(self, event):
        super().paintEvent(event)
        painter = QPainter(self.viewport())
        for item in self.selectedItems():
            rect = self.visualItemRect(item)
            if self.mode:
                handle_rect = QRect(rect.right() - self.fill_handle_size, rect.bottom() - self.fill_handle_size, self.fill_handle_size, self.fill_handle_size)
                painter.fillRect(handle_rect, QColor(250, 250, 250))
        if self.drag_rect:
            painter.setPen(QColor(0, 0, 255))
            painter.drawRect(self.drag_rect)

        pen_color = QColor(255, 255, 255)  # 흰색 경계선
        painter.setPen(pen_color)

        # 첫 번째 열 경계선 그리기
        for row in range(self.rowCount()):
            rect = self.visualRect(self.model().index(row, 0))
            painter.drawLine(rect.topLeft(), rect.bottomLeft())

        painter.end()

    def keyPressEvent(self, event):
        if event.matches(QKeySequence.Copy):
            self.copySelection()
        elif event.matches(QKeySequence.Paste):
            self.pasteSelection()
        elif event.key() == Qt.Key_Delete:
            self.deleteSelection()
        elif event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:
            if self.state() != QAbstractItemView.EditingState:
                # 딜레이를 주어 편집 모드로 진입
                QTimer.singleShot(100, lambda: self.editItem(self.currentItem()))
            else:
                self.moveToNextCell() 
        else:
            super().keyPressEvent(event)

    def copySelection(self):
        selection = self.selectedRanges()
        if selection:
            rows = sorted(range(selection[0].topRow(), selection[0].bottomRow() + 1))
            columns = sorted(range(selection[0].leftColumn(), selection[0].rightColumn() + 1))
            rowcount = len(rows)
            colcount = len(columns)
            table = [[''] * colcount for _ in range(rowcount)]
            for i, row in enumerate(rows):
                for j, col in enumerate(columns):
                    item = self.item(row, col)
                    table[i][j] = item.text() if item else ''
            stream = '\n'.join(['\t'.join(row) for row in table])
            QApplication.clipboard().setText(stream)

    def pasteSelection(self):
        clipboard = QApplication.clipboard()
        if clipboard.mimeData().hasText():
            text = clipboard.text()
            rows = text.split('\n')
            selected = self.selectedRanges()
            if selected:
                start_row = selected[0].topRow()
                start_col = selected[0].leftColumn()
                for i, row in enumerate(rows):
                    columns = row.split('\t')
                    for j, column in enumerate(columns):
                        if start_row + i < self.rowCount() and start_col + j < self.columnCount():
                            self.setCellItemAligned(start_row + i, start_col + j, column)

    def setCellItemAligned(self, row, column, value, align=Qt.AlignCenter):
        item = QTableWidgetItem(str(value))
        item.setTextAlignment(align)
        self.setItem(row, column, item)
                
    def deleteSelection(self):
        for item in self.selectedItems():
            item.setText("")

    def moveToNextCell(self):
        current = self.currentItem()
        if current:
            row = current.row()
            col = current.column()
            if row + 1 < self.rowCount():
                next_item = self.item(row+1, col)
            else:
                next_item = self.item(0, col+1)

            if next_item:
                self.setCurrentItem(next_item)
                 # 딜레이를 주어 편집 모드로 진입
                QTimer.singleShot(100, lambda: self.editItem(self.currentItem()))
    
    def get_column_header(self) -> list:
        return [self.horizontalHeaderItem(i).text() for i in range(self.columnCount())]

    def set_column_value(self, columnname, value):
        headers = self.get_column_header()
        col = headers.index(columnname)
        for row in range(self.rowCount()):
            self.setCellItemAligned(row, col, value)

    def hide_columns(self, column_names):
        """주어진 컬럼명을 기반으로 컬럼을 숨기는 함수"""
        for col in range(self.columnCount()):
            header_item = self.horizontalHeaderItem(col)
            if header_item and header_item.text() in column_names:
                self.setColumnHidden(col, True)  # 컬럼 숨기기
    
    def show_all_columns(self):
        """모든 숨겨진 컬럼을 표시하는 함수"""
        for col in range(self.columnCount()):
            self.setColumnHidden(col, False)  # 모든 컬럼 숨김 해제

    def alignAllCellsCenter(self):
        for row in range(self.rowCount()):
            for col in range(self.columnCount()):
                item = QTableWidgetItem()
                item.setTextAlignment(Qt.AlignCenter)
                self.setItem(row, col, item)

    def set_forecolor_cells(self, rows, columns=None, color="green"):
        """
        Set the text color (foreground) of specific cells in a QTableWidget.
        If columns is None, the entire row will be colored.

        :param table_widget: The QTableWidget instance
        :param rows: List of row indices
        :param columns: List of column indices or None to color entire rows
        :param color: QColor or string representing the color (e.g., 'green')
        """
        for row in rows:
            if columns is None:  # If no specific columns are provided
                columns = range(self.columnCount())  # Iterate over all columns in the row
            for col in columns:
                item = self.item(row, col)  # Get the item at the specified cell
                if item:  # Ensure the item exists
                    item.setForeground(QColor(color))  # Set the text color

    def sort_table_widget(self, primary_column, secondary_column, ascending=True):
        """
        Sort a QTableWidget by multiple columns (primary and secondary).

        :param table_widget: The QTableWidget instance
        :param primary_column: The primary column index to sort by
        :param secondary_column: The secondary column index to sort by
        :param ascending: Sort order (True for ascending, False for descending)
        """
        # Extract data from QTableWidget
        rows = []
        for row in range(self.rowCount()):
            row_data = []
            for col in range(self.columnCount()):
                item = self.item(row, col)
                row_data.append(item.text() if item else "")  # Extract text, or empty string for empty cells
            rows.append(row_data)

        # Sort data using Python's sorted() with a lambda for multi-key sort
        rows_sorted = sorted(
            rows,
            key=lambda x: (x[primary_column], x[secondary_column]),
            reverse=not ascending
        )

        # Clear and repopulate the table with sorted data
        self.clearContents()
        for row_idx, row_data in enumerate(rows_sorted):
            for col_idx, value in enumerate(row_data):
                item = QTableWidgetItem(value)
                item.setTextAlignment(Qt.AlignCenter)
                self.setItem(row_idx, col_idx, item)


class AutoResizeDelegate(QStyledItemDelegate):
    def __init__(self, parent=None):
        super(AutoResizeDelegate, self).__init__(parent)

    def paint(self, painter, option, index):
        option.font.setPointSize(self.calculateFontSize(option, index))
        super(AutoResizeDelegate, self).paint(painter, option, index)

    def calculateFontSize(self, option, index):
        text = index.data()
        if not text:
            return option.font.pointSize()

        base_size = 10
        font = option.font
        font.setPointSize(base_size)
        metrics = QFontMetrics(font)

        while metrics.horizontalAdvance(text) > option.rect.width() or metrics.height() > option.rect.height():
            base_size -= 1
            font.setPointSize(base_size)
            metrics = QFontMetrics(font)
            if base_size <= 1:
                break

        return base_size

class CircularProgress(QWidget):
    def __init__(self):
        super().__init__()

        self.value =0
        self.width=200
        self.height=200
        self.progress_width=10
        self.progress_rounded_cap=True
        self.max_value=100
        self.progress_color= 0xff79c9
        self.enable_text=True
        self.font_family='Segoe UI'
        self.font_size=12
        self.suffix='%'
        self.text_color=0xff79c9
        self.enable_bg=True
        self.bg_color=0x44475a
        self.resize(self.width, self.height)
    
    def add_shadow(self, enable):
        if enable:
            self.shadow = QGraphicsDropShadowEffect(self)
            self.shadow.setBlurRadius(15)  
            self.shadow.setXOffset(0)  
            self.shadow.setYOffset(0)  
            self.shadow.setColor(QColor(0, 0, 0, 80))  
            self.setGraphicsEffect(self.shadow)

    def set_value(self, value):
        self.value=value
        self.repaint()  

    def paintEvent(self, e):
        width = self.width - self.progress_width
        height = self.height - self.progress_width
        margin = self.progress_width / 2
        value = self.value * 360 / self.max_value

        paint = QPainter()
        paint.begin(self)
        paint.setRenderHint(QPainter.Antialiasing)  
        paint.setFont(QFont(self.font_family, self.font_size))  

        rect=QRect(0, 0, self.width, self.height)
        paint.setPen(Qt.NoPen)  
        paint.drawRect(rect)  

        pen = QPen()
        pen.setColor(QColor(self.progress_color))
        pen.setWidth(self.progress_width)

        if self.progress_rounded_cap:
            pen.setCapStyle(Qt.RoundCap)

        paint.setPen(pen)  
        paint.drawArc(margin, margin, width, height, -90 * 16, -value * 16)  

        pen.setColor(QColor(self.text_color))  
        paint.setPen(pen)
        paint.drawText(rect, Qt.AlignCenter, f"{self.value}{self.suffix}")  
        
        paint.end()

class Splashscreen(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_SplashScreen()
        self.ui.setupUi(self)

        self.setWindowFlag(Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)

        self.ui.loading.setStyleSheet("color:#fff")

        self.progress = CircularProgress()
        self.progress.width = 300
        self.progress.height = 300
        self.progress.value=50
        self.progress.font_size =40
        self.progress.setFixedSize(self.progress.width, self.progress.height)

        self.progress.add_shadow(True)
        self.progress.bg_color = QColor(68, 71, 90, 140)

        self.progress.setParent(self.ui.centralwidget)
        self.progress.show()
        self.shadow = QGraphicsDropShadowEffect(self)
        self.shadow.setBlurRadius(15)
        self.shadow.setXOffset(0)
        self.shadow.setYOffset(0)
        self.shadow.setColor(QColor(0, 0, 0, 80))
        self.setGraphicsEffect(self.shadow)

        self.timer = QTimer()
        self.timer.timeout.connect(self.update)
        self.timer.start(20)
        self.show()
    
    def update(self):
        global splash_timer_value
        self.progress.set_value(splash_timer_value)
        if splash_timer_value >= 100:
            self.timer.stop()
            self.main =  CcpManager()
            self.main.setStyleSheet(self.main.get_stylesheet_from_resource(':resources/styles/main_ui.qss'))
            self.close()
        splash_timer_value += 1


class CcpManager(QMainWindow):

    HEADER_LABELS = ['점번호', 'X', 'Y', '도선등급', '도선명', '표지재질', '토지소재(동리)', 
                      '토지소재(지번)', '지적(임야)도', '설치년월일', '조사년월일', '조사자(직)', 
                      '조사자(성명)', '조사내용', '경위도(B)', '경위도(L)', '원점', '표고', '사진파일(경로)', '사진파일명', '위성사진']
    RTK_HEADERS = ['번호', '시작', '종료', '에포크', '수평', '수직', '위도', '경도', '타원체고', 'X', 'Y', 'Z', '지오이드고',
                   'PDOP', 'HDOP', 'VDOP', '장비', '위성수', '솔루션', '사진', '재질', '토지소재(동리)', '토지소재(지번)', '지적(임야)도']
    TEMPLATE = ':resources/templates/template.xlsx'
    TEMPLATE2 = ':resources/templates/template2.xlsx'
    RTK_TEMPLATE = ':resources/templates/RTK_TEMPLATE.xlsx'
    

    def __init__(self):
        super().__init__()
        self._image_folder = None       # 그림파일 폴더
        self.image_extension = ".jpg"   # 그림파일 디폴트 확장자
        self.is_same_name = False       # 도근번호와 그림파일명이 같은가?
        self.rtk_data_file = None       # rtk_data 파일
        self.rtk_data_path = None       # rtk_data 파일폴더
        self.mode = "edit-table"        # 모드 ['edit-table', 'edit-image']
        self.settings = EnvironmentManager()
        self.__width = 400
        self.__minimum_sidebar_width = 250
        self.tr = None                  # tr.dat 파일 경로
        self.setWindowTitle("지적삼각보조(도근)점 성과표 작성")
        self.setWindowIcon(QIcon(':resources/icons/ruler_binder.ico'))
        self.add_toolbar()
        self.setupUi()
        self.showMaximized()
        # self.add_menubar()
    
        self.reflect_settings()
        self.show()

    def setupUi(self):
        self.button_group = QButtonGroup(self)
        # QDockWidget
        side_container = QWidget(self)
        side_layout_main = QVBoxLayout(side_container)
        side_layout_main.setContentsMargins(0, 0, 0, 0)
        temp_label = QLabel(side_container)
        temp_label.setObjectName("temp_label")
        temp_label.setFixedHeight(30)
        side_layout_main.addWidget(temp_label)
        side_layout_detail = QVBoxLayout()
        side_layout_detail.setContentsMargins(10, 10, 10, 10)
        side_layout_detail.addItem(QSpacerItem(10, 10, QSizePolicy.Minimum, QSizePolicy.Expanding))

        # RTK 관측부 ==========
        vlayout_rtk = QVBoxLayout()
        self.input_rtkdata_button = QPushButton(QIcon(":resources/icons/satellite.svg"), f'  {"RTK 기록부 외":^10}', side_container)
        self.input_rtkdata_button.setIconSize(QSize(32, 32))
        self.input_rtkdata_button.setObjectName("input_rtkdata_button")
        self.input_rtkdata_button.setCheckable(True)
        vlayout_rtk.addWidget(self.input_rtkdata_button)
        self.button_group.addButton(self.input_rtkdata_button)
        
        # RTK 관측부 - 서브
        self.rtk_data_sub = QWidget(side_container)
        self.rtk_data_sub.setFixedHeight(440)
        rtk_data_sub_layout = QVBoxLayout()
        # rtk_data_sub_layout.setSpacing(15)
        self.rtk_xlsx_button = QPushButton(QIcon(':resources/icons/xlsx-file.svg'), f'  {"관측데이터(xlsx)":^14} ', side_container)
        self.rtk_xlsx_button.setIconSize(QSize(24, 24))
        self.rtk_sort_button = QPushButton(QIcon(':resources/icons/sort-from-top-to-bottom.svg'), f' {"번호-시작 정렬":^14}', side_container)
        self.rtk_sort_button.setIconSize(QSize(24, 24))
        self.rtk_timecheck_button=QPushButton(QIcon(':resources/icons/clock-circle.svg'), f'   {"타임 조정":^14}   ', side_container)
        self.rtk_timecheck_button.setIconSize(QSize(24, 24))
        self.rtk_cif_button = QPushButton(QIcon(':resources/icons/cif-file.svg'), f'   {"Cif/Shp 입력":^14}    ', side_container)
        self.rtk_cif_button.setIconSize(QSize(24, 24))

        
        self.surveyor_name = QComboBox(side_container)  # 관측자(명)
        self.surveyor_name.setObjectName("surveyor_name")
        hlayout_rtk1 = QHBoxLayout()

        self.surveyor_grade = QComboBox(side_container) # 관측자(직)
        self.surveyor_grade.setObjectName("surveyor_grade")
        self.surveyor_grade.addItems([f"국토정보직 {str(i)}급" for i in range(1, 8)])
        self.surveyor_grade.setCurrentIndex(3)
        self.survey_license = QComboBox(side_container)  # 관측자 자격
        self.survey_license.setObjectName("survey_license")
        self.survey_license.addItems(["지적기술사", "지적기사", "지적산업기사", "지적기능사"])
        self.survey_license.setCurrentIndex(1)
        self.survey_license.setFixedWidth(90)

        hlayout_rtk1.addWidget(self.surveyor_grade)
        hlayout_rtk1.addWidget(self.survey_license)

        hlayout_rtk2 = QHBoxLayout()
        self.reception_number = QLineEdit(side_container)  # 접수번호
        self.ccp_type = QComboBox(side_container)  # 기준점 종류
        self.ccp_type.setObjectName("ccp_type")
        self.ccp_type.addItems(["지적삼각점", "지적삼각보조점", "지적도근점"])
        self.ccp_type.setCurrentIndex(2)
        self.ccp_type.setFixedWidth(90)

        hlayout_rtk2.addWidget(self.reception_number)
        hlayout_rtk2.addWidget(self.ccp_type)

        self.ref_jibun = QLineEdit(side_container)  # 대표 지번
        self.jigu_name = QLineEdit(side_container)  # 지구명
        self.jigu_attr = QLineEdit(side_container)  # 지구특성      
        self.antena_sn = QLineEdit(side_container)  # 안테나 명(번호)        

        self.reception_number.setPlaceholderText("접수번호")
        self.ref_jibun.setPlaceholderText("대표 지번(용인시 처인구...)")
        self.jigu_name.setPlaceholderText("지구명")
        self.jigu_attr.setPlaceholderText("지구특성")
        self.antena_sn.setPlaceholderText("안테나 명(번호)")
        self.surveyor_grade.setPlaceholderText("관측자(직)")
        self.surveyor_name.setPlaceholderText("관측자(성명)")

        rtk_data_sub_layout.addWidget(self.rtk_xlsx_button)
        rtk_data_sub_layout.addWidget(self.rtk_sort_button)
        rtk_data_sub_layout.addWidget(self.rtk_timecheck_button)
        rtk_data_sub_layout.addWidget(self.rtk_cif_button)
        rtk_data_sub_layout.addWidget(self.surveyor_name)
        rtk_data_sub_layout.addLayout(hlayout_rtk1)
        rtk_data_sub_layout.addLayout(hlayout_rtk2)
        rtk_data_sub_layout.addWidget(self.ref_jibun)
        rtk_data_sub_layout.addWidget(self.jigu_name)
        rtk_data_sub_layout.addWidget(self.jigu_attr)
        rtk_data_sub_layout.addWidget(self.antena_sn)

        self.rtk_record_button = QPushButton(QIcon(':resources/icons/document-add.svg'), f'  {"관측기록부 등":^10}  ', side_container)
        self.rtk_record_button.setIconSize(QSize(24, 24))

        rtk_data_sub_layout.addWidget(self.rtk_record_button)

        rtk_hlayout = QHBoxLayout()
        rtk_hspacer = QSpacerItem(10, 10, QSizePolicy.Expanding, QSizePolicy.Minimum)
        self.rtk_apply_button = QPushButton(QIcon(':resources/icons/cpu.svg'), f'  {"데이터전환"}  ', side_container)
        self.rtk_apply_button.setIconSize(QSize(24, 24))

        rtk_hlayout.addItem(rtk_hspacer)
        rtk_hlayout.addWidget(self.rtk_apply_button)
        rtk_data_sub_layout.addLayout(rtk_hlayout)

        rtk_data_sub_layout.addItem(QSpacerItem(10, 10, QSizePolicy.Minimum, QSizePolicy.Expanding))
        self.rtk_data_sub.setLayout(rtk_data_sub_layout)
        self.rtk_data_sub.setVisible(False)
        vlayout_rtk.addWidget(self.rtk_data_sub)
        vlayout_rtk.setSpacing(5)        
        side_layout_detail.addLayout(vlayout_rtk)

        # 데이터 입력 ==========
        vlayout_data = QVBoxLayout()
        self.input_data_button = QPushButton(QIcon(':resources/icons/file-text.svg'), f'  {"데이터 입력":^10}', side_container)  
        self.input_data_button.setIconSize(QSize(32, 32))
        self.input_data_button.setObjectName("input_data_button")
        self.input_data_button.setCheckable(True)
        vlayout_data.addWidget(self.input_data_button)
        self.button_group.addButton(self.input_data_button)
        
        # 데이터 입력 - 서브
        self.input_data_sub = QWidget(side_container)
        self.input_data_sub.setFixedHeight(100)
        input_data_sub_layout = QVBoxLayout()
        input_data_sub_layout.setSpacing(15)
        self.tr_dat_button = QPushButton(QIcon(':resources/icons/target.svg'), f'    {"tr.dat 입력":^10}     ', side_container)        
        self.tr_dat_button.setIconSize(QSize(24, 24))
        self.load_project_button = QPushButton(QIcon(':resources/icons/file-right.svg'), f' {"기존 프로젝트":^10}  ', side_container)
        self.load_project_button.setIconSize(QSize(24, 24))
        input_data_sub_layout.addWidget(self.tr_dat_button)
        input_data_sub_layout.addWidget(self.load_project_button)
        input_data_sub_layout.addItem(QSpacerItem(10, 10, QSizePolicy.Minimum, QSizePolicy.Expanding))
        self.input_data_sub.setLayout(input_data_sub_layout)
        self.input_data_sub.setVisible(False)
        vlayout_data.addWidget(self.input_data_sub)
        vlayout_data.setSpacing(5)
        side_layout_detail.addLayout(vlayout_data)
        
        # 공통값 입력 ==========
        vlayout_common = QVBoxLayout()
        self.common_input_button = QPushButton(QIcon(':resources/icons/bill-list.svg'), f'  {"공통값 입력":^10}', side_container)
        self.common_input_button.setIconSize(QSize(32, 32))
        self.common_input_button.setObjectName("common_input_button")
        self.common_input_button.setCheckable(True)
        vlayout_common.addWidget(self.common_input_button)
        self.button_group.addButton(self.common_input_button)

        # 공통값 입력 - 서브
        common_input_sub = QWidget(side_container)
        common_input_sub_layout = QVBoxLayout()
        common_input_sub.setFixedHeight(300)

        self.grade_input = QLineEdit(side_container)
        self.name_input = QLineEdit(side_container)
        self.install_date_input = QLineEdit(side_container)
        self.survey_date_input = QLineEdit(side_container)
        self.surveyor_position_input = QLineEdit(side_container)
        self.surveyor_input = QLineEdit(side_container)
        self.findings_input = QLineEdit(side_container)
        self.origin_input = QLineEdit(side_container)

        self.grade_input.setPlaceholderText('도선등급 ex) 1')
        self.name_input.setPlaceholderText('도선명 ex) 가')
        self.install_date_input.setPlaceholderText('설치년월일 ex) 2020-10-12')
        self.survey_date_input.setPlaceholderText('조사년월일 ex) 2021-02-23')
        self.surveyor_position_input.setPlaceholderText('조사자(직)')
        self.surveyor_input.setPlaceholderText('조사자')
        self.findings_input.setPlaceholderText('조사내용')
        self.findings_input.setText("신설")
        self.origin_input.setPlaceholderText('원점')
        self.origin_input.setText('세계측지계(중부)')

        common_input_sub_layout.addWidget(self.grade_input)
        common_input_sub_layout.addWidget(self.name_input)
        common_input_sub_layout.addWidget(self.install_date_input)
        common_input_sub_layout.addWidget(self.survey_date_input)
        common_input_sub_layout.addWidget(self.surveyor_position_input)
        common_input_sub_layout.addWidget(self.surveyor_input)
        common_input_sub_layout.addWidget(self.findings_input)
        common_input_sub_layout.addWidget(self.origin_input)

        hlayout1 = QHBoxLayout()
        hspacer1 = QSpacerItem(10, 10, QSizePolicy.Expanding, QSizePolicy.Minimum)
        self.common_apply_button = QPushButton(QIcon(':resources/icons/cpu.svg'), '  적용', side_container)
        self.common_apply_button.setIconSize(QSize(24, 24))
        hlayout1.addItem(hspacer1)
        hlayout1.addWidget(self.common_apply_button)
        common_input_sub_layout.addLayout(hlayout1)
        common_input_sub_layout.addItem(QSpacerItem(10, 10, QSizePolicy.Minimum, QSizePolicy.Expanding))
        common_input_sub.setLayout(common_input_sub_layout)
        
        common_input_sub.setVisible(False)
        vlayout_common.addWidget(common_input_sub)
        vlayout_common.setSpacing(5)
        side_layout_detail.addLayout(vlayout_common)

        # 토지소재지 입력 ==================
        vlayout_land = QVBoxLayout()
        self.land_data_button = QPushButton(QIcon(':resources/icons/streets-map-point.svg'), f'  {"소재지 검색":^10}', side_container)
        self.land_data_button.setIconSize(QSize(32, 32))
        self.land_data_button.setObjectName("land_data_button")
        self.land_data_button.setCheckable(True)
        vlayout_land.addWidget(self.land_data_button)
        self.button_group.addButton(self.land_data_button)
        
        # 토지소재지 입력 - 서브
        land_data_sub = QWidget(side_container)
        land_data_sub.setFixedHeight(100)
        land_data_sub_layout = QVBoxLayout()
        land_data_sub_layout.setSpacing(15)
        self.cif_button = QPushButton(QIcon(':resources/icons/cif-file.svg'), '   Cif에서 검색', side_container)
        self.cif_button.setIconSize(QSize(24, 24))        
        self.shp_button = QPushButton(QIcon(':resources/icons/shape-file.svg'), '  Shp에서 검색', side_container)
        self.shp_button.setIconSize(QSize(24, 24))
        land_data_sub_layout.addWidget(self.cif_button)
        land_data_sub_layout.addWidget(self.shp_button)
        land_data_sub_layout.addItem(QSpacerItem(10, 10, QSizePolicy.Minimum, QSizePolicy.Expanding))
        land_data_sub.setLayout(land_data_sub_layout)
        
        land_data_sub.setVisible(False)
        vlayout_land.addWidget(land_data_sub)
        vlayout_land.setSpacing(5)
        side_layout_detail.addLayout(vlayout_land)

        # 사진관리 ================
        vlayout_image = QVBoxLayout()
        self.image_management_button = QPushButton(QIcon(':resources/icons/gallery-wide.svg'), f'  {"이미지 삽입":^10}', side_container)
        self.image_management_button.setIconSize(QSize(32, 32))
        self.image_management_button.setObjectName("image_management_button")
        self.image_management_button.setCheckable(True)
        vlayout_image.addWidget(self.image_management_button)
        self.button_group.addButton(self.image_management_button)

        # 사진관리 - 서브
        image_management_sub = QWidget(side_container)
        image_management_sub.setFixedHeight(230)
        image_management_sub_layout = QVBoxLayout()

        self.get_image_button = QPushButton(QIcon(':resources/icons/album.svg'), '  사진폴더 선택', side_container)
        self.get_image_button.setIconSize(QSize(24, 24))
        extension_group = QGroupBox(side_container)
        extension_group.setTitle('확장자')
        vlayout1 = QVBoxLayout()
        self.jpg_radio = QRadioButton('.jpg', extension_group)
        self.jpg_radio.setChecked(True)
        self.jpeg_radio = QRadioButton('.jpeg', extension_group)
        self.png_radio = QRadioButton('.png', extension_group)
        vlayout1.addWidget(self.jpg_radio)
        vlayout1.addWidget(self.jpeg_radio)
        vlayout1.addWidget(self.png_radio)
        extension_group.setLayout(vlayout1)
        self.image_edited_check = QCheckBox(side_container)
        self.image_edited_check.setText('  편집 이미지(_edit)')
        hlayout2 = QHBoxLayout()
        hspacer2 = QSpacerItem(10, 10, QSizePolicy.Expanding, QSizePolicy.Minimum)
        self.image_apply_button = QPushButton(QIcon(':resources/icons/cpu.svg'), '  적용', side_container)
        self.image_apply_button.setIconSize(QSize(24, 24))
        hlayout2.addItem(hspacer2)
        hlayout2.addWidget(self.image_apply_button)
        
        image_management_sub_layout.addWidget(self.get_image_button)
        image_management_sub_layout.addWidget(extension_group)
        image_management_sub_layout.addWidget(self.image_edited_check)
        image_management_sub_layout.addLayout(hlayout2)
        image_management_sub_layout.addItem(QSpacerItem(10, 10, QSizePolicy.Minimum, QSizePolicy.Expanding))
        image_management_sub.setLayout(image_management_sub_layout)
        image_management_sub.setVisible(False)
        vlayout_image.addWidget(image_management_sub)
        vlayout_image.setSpacing(5)
        side_layout_detail.addLayout(vlayout_image)

        # 내보내기 ==================
        vlayout_export = QVBoxLayout()
        self.export_button = QPushButton(QIcon(':resources/icons/ssd-round.svg'), f'   {"내보내기":^10} ', side_container)
        self.export_button.setIconSize(QSize(32, 32))
        self.export_button.setObjectName("export_button")
        self.export_button.setCheckable(True)
        vlayout_export.addWidget(self.export_button)
        self.button_group.addButton(self.export_button)
        
        # 내보내기 - 서브
        export_data_sub = QWidget(side_container)
        export_data_sub.setFixedHeight(100)
        export_data_sub_layout = QVBoxLayout()
        export_data_sub_layout.setSpacing(15)
        self.export_project_button = QPushButton(QIcon(':resources/icons/file-left.svg'), '   프로젝트 저장 ', side_container)
        self.export_project_button.setIconSize(QSize(24, 24))
        self.export_xlsx_button = QPushButton(QIcon(':resources/icons/xlsx-file.svg'), '  성과표 엑셀저장', side_container)
        self.export_xlsx_button.setIconSize(QSize(24, 24))
        export_data_sub_layout.addWidget(self.export_project_button)
        export_data_sub_layout.addWidget(self.export_xlsx_button)
        export_data_sub_layout.addItem(QSpacerItem(10, 10, QSizePolicy.Minimum, QSizePolicy.Expanding))
        export_data_sub.setLayout(export_data_sub_layout)
        export_data_sub.setVisible(False)
        vlayout_export.addWidget(export_data_sub)
        vlayout_export.setSpacing(5)        
        side_layout_detail.addLayout(vlayout_export)

        side_layout_detail.addItem(QSpacerItem(10, 150, QSizePolicy.Minimum, QSizePolicy.Expanding))

        # 기타 툴 ==================
        vlayout_extra = QVBoxLayout()
        self.extra_tools_button = QPushButton(QIcon(':resources/icons/settings.svg'), f'  {"기타 툴들":^10} ', side_container)
        self.extra_tools_button.setIconSize(QSize(32, 32))
        self.extra_tools_button.setObjectName("extra_tools")
        self.extra_tools_button.setCheckable(True)
        vlayout_extra.addWidget(self.extra_tools_button)
        self.button_group.addButton(self.extra_tools_button)
        
        # 기타 툴 - 서브
        extra_tools_sub = QWidget(side_container)
        extra_tools_sub.setFixedHeight(140)
        extra_tools_sub_layout = QVBoxLayout()
        extra_tools_sub_layout.setSpacing(15)
        self.update_code_button = QPushButton(QIcon(':resources/icons/server-square-update.svg'), '법정동코드 업데이트', side_container)
        self.update_code_button.setIconSize(QSize(24, 24))
        self.classify_image_button = QPushButton(QIcon(':resources/icons/gallery-edit.svg'), '  사진파일명 변경   ', side_container)
        self.classify_image_button.setIconSize(QSize(24, 24))
        self.settings_button = QPushButton(QIcon(':resources/icons/settings.svg'), '      관측자 설정     ', side_container)
        self.settings_button.setIconSize(QSize(24, 24))
        extra_tools_sub_layout.addWidget(self.update_code_button)
        extra_tools_sub_layout.addWidget(self.classify_image_button)
        extra_tools_sub_layout.addWidget(self.settings_button)

        extra_tools_sub_layout.addItem(QSpacerItem(10, 10, QSizePolicy.Minimum, QSizePolicy.Expanding))
        extra_tools_sub.setLayout(extra_tools_sub_layout)
        extra_tools_sub.setVisible(False)
        vlayout_extra.addWidget(extra_tools_sub)
        vlayout_extra.setSpacing(5)

        # 사이드바 레이아웃 정리
        side_layout_detail.addLayout(vlayout_extra)
        side_layout_detail.setSpacing(20)
        side_layout_main.addItem(QSpacerItem(10, 10, QSizePolicy.Minimum, QSizePolicy.Expanding))
        side_layout_main.addLayout(side_layout_detail)
        side_layout_main.addItem(QSpacerItem(10, 10, QSizePolicy.Minimum, QSizePolicy.Expanding))

        # 버튼그룹 내의 버튼은 하나만 선택할 수 있게
        self.button_group.setExclusive(True)  
        scroll_area = self.create_scroll_area(side_container, parent=self)
        self.sidemenu = self.add_dockableWidget("테이블 편집", scroll_area)
        self.sidemenu.setFixedWidth(self.__minimum_sidebar_width)
        self.setResizable(self.sidemenu, self.__minimum_sidebar_width)
     
        # 시그널-슬롯 연결
        self.input_rtkdata_button.toggled.connect(self.input_rtkdata_toggle)
        self.input_data_button.toggled.connect(self.input_rtkdata_toggle)
        self.common_input_button.toggled.connect(common_input_sub.setVisible)
        self.image_management_button.toggled.connect(image_management_sub.setVisible)
        self.land_data_button.toggled.connect(land_data_sub.setVisible)
        self.export_button.toggled.connect(export_data_sub.setVisible)
        self.extra_tools_button.toggled.connect(extra_tools_sub.setVisible)

        #### main widget #########################################################################################
        self.main_frame = QFrame(self)
        main_frame_layout = QHBoxLayout(self.main_frame)
        main_frame_layout.setContentsMargins(0, 0, 0, 0)  
        main_frame_layout.setSpacing(0)  

        # custom table widget
        vlayout2 = QVBoxLayout()
        vlayout2.setContentsMargins(0, 0, 0, 0)
        self.band = QFrame(self)
        self.band.setObjectName("band")
        hlayout_table = QHBoxLayout(self.band)
        hlayout_table.setContentsMargins(0, 0, 0, 0)
        column_hide_icon = QIcon()
        column_hide_icon.addFile(':resources/icons/column-hide.svg', QSize(), QIcon.Normal, QIcon.Off)
        column_hide_icon.addFile(':resources/icons/column-show.svg', QSize(), QIcon.Normal, QIcon.On)
        self.column_hide = QPushButton(self)
        self.column_hide.setIcon(column_hide_icon)
        self.column_hide.setCheckable(True)
        self.column_hide.setIconSize(QSize(24, 24))
        self.column_hide.setObjectName("column-hide")
        self.column_hide.setToolTip("컬럼 숨기기")
        self.column_hide.setFixedWidth(30)
        self.fill_number = QPushButton(icon=QIcon(':resources/icons/fill-numbers.svg'), parent=self)
        self.fill_number.setIconSize(QSize(24, 24))
        self.fill_number.setObjectName("fill_number")
        self.fill_number.setToolTip("자동번호 입력")
        self.fill_number.setFixedWidth(30)
        self.temp_label1 = QLabel(self)
        self.temp_label1.setObjectName("temp_label1")
        self.table_to_tr = QPushButton(icon=QIcon(':resources/icons/point_black.png'), parent=self)
        self.table_to_tr.setIconSize(QSize(24, 24))
        self.table_to_tr.setObjectName("table_to_tr")
        self.table_to_tr.setToolTip("tr.dat 저장")
        self.table_to_tr.setFixedWidth(30)
        self.save_table = QPushButton(icon=QIcon(':resources/icons/diskette.svg'), parent=self)
        self.save_table.setIconSize(QSize(24, 24))
        self.save_table.setObjectName("save_table")
        self.save_table.setToolTip("변경내용 저장(xlsx)")
        self.save_table.setFixedWidth(30)
        self.band.setFixedHeight(30)
        hlayout_table.addWidget(self.column_hide)
        hlayout_table.addWidget(self.fill_number)
        hlayout_table.addWidget(self.temp_label1)
        hlayout_table.addWidget(self.table_to_tr)
        hlayout_table.addWidget(self.save_table)

        self.notice_frame = QFrame(self)
        self.notice_frame.setFixedWidth(1100)
        notice_frame_layout = QVBoxLayout(self.notice_frame)
        self.text_area = QTextBrowser(self.notice_frame)
        self.text_area.setObjectName("text_area")
        html_text = """
        <style>
        .lawcon {margin-left: 20px; margin-right: 20px; margin-top: 10px; margin-bottom: 10px;}

        p.gtit{
            color:#151594;
            font-size: 2rem; 
            margin:0 0 0 30px;
            padding:10px 0 0 14px;
            font-weight:bold;
        }
        div.subtit3{
            float:right;
            margin:10px 17px 10px 0;
            padding:6px 10px 8px;
            border-top:1px solid #e3e3e3;
            border-bottom:1px solid #e3e3e3;
            background:#f8f8f1;
        }
        p.pty1_de2_1{
            margin: 5px 0 0 67px;
        }
        p.pty1_de2h{
            margin: 5px 0 0 100px;
        }
        span.bl{
            color:#151594;
            font-weight:bold;
        }
        h2{
            margin:20px 0 10px 0;
            color:#1C1C1C; 
            text-align:center;
            position: relative;
        }
        div.subtit1 {
            color:#3c7fbc; 
            text-align:center;
            padding: 0px 0px 13px 0px; /*2024.01.11. #33957 수정*/
        }

        div.pty1_table {
            text-align: center;
        }

        td, th {
            padding: 5px;
            text-align: center;
        }
        </style>
        <div>  &nbsp; </div>
        <div id="conTop"> 
            <h2>GNSS에 의한 지적측량규정</h2> 
            <div class="subtit1">
                [시행 2020. 8. 10.] [국토교통부예규 제304호, 2020. 8. 10., 일부개정] 
            </div> 
        </div>

        <p class="gtit" style="padding-top: 5px; margin-bottom: 0px;"><h3>제2장 선점 및 측량</h3></p>
        <div class="lawcon"> 
        <p class="pty1_p4"><span class="bl">제6조(관측)</span> </span> ① 관측 시 위성의 조건은 다음 각 호의 기준에 의한다.&nbsp; 
        <p class="pty1_de2h">1. 관측점으로부터 위성에 대한 고도각이 15°이상에 위치할 것&nbsp;</p> 
        <p class="pty1_de2h">2. 위성의 작동상태가 정상일 것&nbsp;</p> 
        <p class="pty1_de2h">3. 관측점에서 동시에 수신 가능한 위성 수는 정지측량에 의하는 경우에는 4개 이상, 이동측량에 의하는 경우에는 5개 이상일 것&nbsp;</p> 
        <p class="pty1_de2_1">② GNSS측량기에 입력하는 안테나의 높이 등에 관하여는 GNSS측량기에서 정해진 방법에 따라 측정하고, 관측 후 확인한다.&nbsp;</p> 
        <p class="pty1_de2_1">③ 관측 시 주의사항은 다음 각 호와 같다.&nbsp;</p> 
        <p class="pty1_de2h">1. 안테나 주위의 10미터이내에는 자동차 등의 접근을 피할 것&nbsp;</p> 
        <p class="pty1_de2h">2. 관측 중에는 무전기 등 전파발신기의 사용을 금한다. 다만, 부득이한 경우에는 안테나로부터 100미터이상의 거리에서 사용할 것&nbsp;</p> 
        <p class="pty1_de2h">3. 발전기를 사용하는 경우에는 안테나로부터 20미터 이상 떨어진 곳에서 사용할 것&nbsp;</p> 
        <p class="pty1_de2h">4. 관측 중에는 수신기 표시장치 등을 통하여 관측상태를 수시로 확인하고 이상 발생시에는 재관측을 실시할 것&nbsp;</p> 
        <!-- 호 표시 end --> 
        <p class="pty1_de2_1">④ 관측 완료 후 점검결과 제1항 내지 제3항의 관측조건에 맞지 아니한 경우에는 다시 관측을 하여야 한다.&nbsp;</p> 

        <p class="pty1_de2_1">⑤ 지적위성측량을 실시하는 경우에는 지적위성측량관측부를 작성하여야 한다.&nbsp;</p> 
        </div>
        <p> </p>

        <div class="lawcon"> 
        <p class="pty1_p4"><span class="bl">제7조(정지측량)</span> GNSS측량기를 사용하여 정지측량방법으로 기초측량 또는 세부측량을 하고자 하는 때에는 다음 각 호의 기준에 의한다. &nbsp; 
        <p class="pty1_de2h">1. 기지점과 소구점에 GNSS측량기를 동시에 설치하여 세션단위로 실시할 것&nbsp;</p> 
        <p class="pty1_de2h">2. 관측성과의 기선벡터 점검을 위하여 다른 세션에 속하는 관측망과 1변 이상이 중복되게 관측할 것.&nbsp;</p> 
        <p class="pty1_de2h">3. 관측시간 등은 다음 표에 의할 것<br>&nbsp;</p> 
        <div class="pty1_table">
            <table border="1" cellpadding="0" cellspacing="0" width="80%">
                <tr>
                    <th>구분</th> 
                    <th>지적삼각측량</th>
                    <th>지적삼각보조측량</th> 
                    <th>지적도근측량</th>
                    <th>세부측량</th>
                </tr>
                <tr> 
                    <td>기지점과의 거리</td> 
                    <td>10Km 미만</td> 
                    <td>5Km 미만</td>
                    <td>2Km 미만</td>
                    <td>1Km 미만</td>
                </tr>
                <tr> 
                    <td>세션 관측시간</td> 
                    <td>60분 이상</td> 
                    <td>30분 이상</td>
                    <td>10분 이상</td>
                    <td>5분 이상</td> 
                </tr> 
                <tr> 
                    <td>데이터 취득간격</td> 
                    <td>30초 이하</td> 
                    <td>30초 이하</td> 
                    <td>15초 이하</td>
                    <td>15초 이하</td>
                </tr>
            </table>
        </div>
      </div>
      <p> </p>
      <div class="lawcon"> 
       <p class="pty1_p4"><span class="bl">제8조(이동측량)</span>  ① GNSS측량기를 사용하여 지적도근측량 또는 세부측량을 하고자 하는 경우에는 단일기준국 실시간 이동측량 또는 다중기준국 실시간 이동측량에 의한다.&nbsp; 
       <p class="pty1_de2_1">② 단일기준국 실시간 이동측량(Single-RTK) 및 다중기준국 실시간 이동측량(Network-RTK)으로 실시할 경우 기준은 다음 각 호와 같다.&nbsp;</p> 
       <p class="pty1_de2h">1. 관측전 이동국 GNSS측량기의 초기화 작업을 완료할 것&nbsp;</p> 
       <p class="pty1_de2h">2. 관측 중 위성신호의 단절 또는 통신장치의 이상으로 보정정보를 안정적으로 수신할 수 없는 경우 이동국 GNSS측량기를 재초기화 할 것&nbsp;</p> 
       <p class="pty1_de2h">3. GNSS측량기 안테나를 기준으로 고도각 15°이상에 정상 작동중인 GNSS위성이 5개 이상일 것&nbsp;</p> 
       <p class="pty1_de2h">4. GNSS측량기에 표시하는 PDOP이 3이상이거나 위치정밀도가 수평 ±3㎝ 이상 또는 수직 ±5㎝ 이상인 경우 관측을 중지할 것&nbsp;</p> 
       <p class="pty1_de2h">5. 1, 2회의 관측치가 제5항제4호의 오차 이내일 경우에는 1회 관측치를 기준으로 결과부를 작성&nbsp;</p> 
       <p class="pty1_de2h">6. 지역좌표를 구하고자 할 경우에는 GNSS측량기에서 제공하는 소프트웨어를 이용하여 좌표변환 계산방법에 의할 것&nbsp;</p> 
       <p class="pty1_de2h">7. 관측시간 및 관측횟수는 다음 표에 따른다. 다만, 단일기준국 실시간 이동측량(Single-RTK 측량) 시 기선거리는 5㎞이내로 한다.<br>&nbsp;</p> 
       <div class="pty1_table">
            <table border="1" cellpadding="0" cellspacing="0" width="80%">
                <tr>
                    <th>구분</th> 
                    <th>관측횟수</th>
                    <th>관측 간격</th> 
                    <th>관측시간(고정해)</th>
                    <th>데이터 쥐득간격</th>
                </tr>
                <tr> 
                    <td>도근측량</td> 
                    <td>2회</td> 
                    <td>60분 이상</td>
                    <td>60초 이상</td>
                    <td>1초</td>
                </tr>
                <tr> 
                    <td>세부측량</td> 
                    <td>2회</td> 
                    <td>60분 이상</td>
                    <td>15초 이상</td>
                    <td>1초</td> 
                </tr> 
            </table>
        </div>
       <p class="pty1_de2_1">③ 단일기준국 실시간 이동측량(Single-RTK 측량)에 의한 방법은 다음 각 호와 같다.&nbsp;</p> 
       <p class="pty1_de2h">1. 기지점에 기준국을 설치하고 위치를 결정하고자 하는 지적도근점 이나 경계점 등을 이동국으로 하여 GNSS측량기를 순차적으로 설치하여 이동하며 관측을 실시할 것&nbsp;</p> 
       <p class="pty1_de2h">2. 관측 노선(단위)을 포함하도록 기준국을 달리하여 2회 관측할 것&nbsp;</p> 
       <p class="pty1_de2_1">④ 다중기준국 실시간 이동측량(Network- RTK 측량)에 의한 방법은 다음 각 호와 같다.&nbsp;</p> 
       <p class="pty1_de2h">1. 이동국은 보정정보 생성에 사용되는 상시관측소 네트워크 내부에 있을 것. 다만 부득이한 경우 네트워크 외부에서 10㎞ 이내일 것.&nbsp;</p> 
       <p class="pty1_de2h">2. 통신장치를 이용하여 위성기준점 네트워크 보정신호를 수신하여 고정해를 얻고 이동국을 순차로 이동하면서 관측을 실시할 것&nbsp;</p> 
       <p class="pty1_de2_1">⑤ 단일기준국 실시간 이동측량(Single-RTK) 및 다중기준국 실시간 이동측량(Network-RTK)에 의한 경우 제2항제1호부터 제4호까지의 조건을 만족하지 못하거나 다음 각 호의 경우에는 측량방법을 달리하여 실시한다.&nbsp;</p> 
       <p class="pty1_de2h">1. 초기화 시간이 3회 이상 3분을 초과하는 경우&nbsp;</p> 
       <p class="pty1_de2h">2. 보정정보의 송수신이 불안정한 경우&nbsp;</p> 
       <p class="pty1_de2h">3. 보정정보 지연시간이 5초 이상인 경우&nbsp;</p> 
       <p class="pty1_de2h">4. 세션 간 측량성과의 오차가 5.0㎝를 초과하는 경우&nbsp;</p> 
       <p></p> 
      </div>
        
       <p></p> 
      </div>        


        """
        self.text_area.setHtml(html_text)
        self.text_area.setStyleSheet('color:black; background-color: #cbcbcb;')
        notice_frame_layout.addWidget(self.text_area)

        self.rtk_table_widget = CustomTableWidget()
        self.rtk_table_widget.setObjectName("rtk_table_widget")
        self.rtk_table_widget.setColumnCount(len(self.RTK_HEADERS))
        self.rtk_table_widget.setRowCount(5)
        self.rtk_table_widget.setHorizontalHeaderLabels(self.RTK_HEADERS)
        self.rtk_table_widget.setWordWrap(False)
        self.rtk_table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.rtk_table_widget.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Preferred)
        self.rtk_table_widget.setColumnWidth(self.rtk_table_widget.get_column_header().index("번호"), 110)
        self.rtk_table_widget.setColumnWidth(self.rtk_table_widget.get_column_header().index("시작"), 150)
        self.rtk_table_widget.setColumnWidth(self.rtk_table_widget.get_column_header().index("종료"), 150)
        self.rtk_table_widget.setColumnWidth(self.rtk_table_widget.get_column_header().index("에포크"), 60)
        self.rtk_table_widget.setColumnWidth(self.rtk_table_widget.get_column_header().index("수평"), 80)
        self.rtk_table_widget.setColumnWidth(self.rtk_table_widget.get_column_header().index("수직"), 80)
        self.rtk_table_widget.setColumnWidth(self.rtk_table_widget.get_column_header().index("위도"), 130)
        self.rtk_table_widget.setColumnWidth(self.rtk_table_widget.get_column_header().index("경도"), 140)
        self.rtk_table_widget.setColumnWidth(self.rtk_table_widget.get_column_header().index("타원체고"), 80)
        self.rtk_table_widget.setColumnWidth(self.rtk_table_widget.get_column_header().index("위성수"), 60)
        self.rtk_table_widget.setColumnWidth(self.rtk_table_widget.get_column_header().index("PDOP"), 50)
        self.rtk_table_widget.setColumnWidth(self.rtk_table_widget.get_column_header().index("HDOP"), 50)
        self.rtk_table_widget.setColumnWidth(self.rtk_table_widget.get_column_header().index("VDOP"), 50)
        self.rtk_table_widget.setColumnWidth(self.rtk_table_widget.get_column_header().index("토지소재(동리)"), 240)   

        self.table_widget = CustomTableWidget(self)
        self.table_widget.setObjectName("table_widget")
        self.table_widget.setColumnCount(len(self.HEADER_LABELS))
        self.table_widget.setRowCount(5)
        self.table_widget.setHorizontalHeaderLabels(self.HEADER_LABELS)
        self.table_widget.setWordWrap(False)
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table_widget.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Preferred)
        self.table_widget.verticalHeader().hide()
        
        vlayout2.addWidget(self.band)
        vlayout2.addWidget(self.notice_frame)
        vlayout2.addWidget(self.rtk_table_widget)
        vlayout2.addWidget(self.table_widget)
        main_frame_layout.addLayout(vlayout2)
 
        # AutoResizeDelegate 설정
        delegate = AutoResizeDelegate(self.table_widget)
        self.table_widget.setItemDelegate(delegate)        
        self.table_widget.setContextMenuPolicy(Qt.NoContextMenu)
        self.alignAllCellsCenter()
        self.table_widget.setColumnWidth(self.table_widget.get_column_header().index("토지소재(동리)"), 200)
        self.table_widget.setColumnWidth(self.table_widget.get_column_header().index("사진파일(경로)"), 350)

        ## image editor widget
        self.image_editor_frame = QFrame(self)
        self.image_editor_frame.setObjectName("image_editor_frame")
        image_editor_frame_layout = QHBoxLayout()
        image_editor_frame_layout.setContentsMargins(0, 0, 0, 0)
        self.image_editor = ImageEditor(self.image_editor_frame)
        self.image_editor.setObjectName("image_editor")
        self.image_editor.setAttribute(Qt.WA_TranslucentBackground)
        image_editor_frame_layout.addWidget(self.image_editor)
        self.image_editor_frame.setLayout(image_editor_frame_layout)
        self.image_editor_frame.hide()
  
        main_frame_layout.addWidget(self.image_editor_frame)
       
        self.setCentralWidget(self.main_frame)
        self.statusbar = QStatusBar(self)
        self.setStatusBar(self.statusbar)
        self.status_message = QLabel(self)
        self.statusbar.addPermanentWidget(self.status_message)

        # signal-slot connection
        self.surveyor_name.currentIndexChanged.connect(self.surveyor_changed)
        self.column_hide.toggled.connect(self.rtk_table_hide_column)
        self.fill_number.clicked.connect(self.auto_fill_number)
        self.table_to_tr.clicked.connect(self.table_to_trdat)
        self.rtk_xlsx_button.clicked.connect(self.loadRTKdata)
        self.save_table.clicked.connect(self.save_rtk_table)
        self.rtk_cif_button.clicked.connect(self.rtk_location)
        self.rtk_sort_button.clicked.connect(self.rtk_sort)
        self.rtk_timecheck_button.clicked.connect(self.rtk_timecheck)
        self.rtk_record_button.clicked.connect(self.rtk_report_all)
        self.rtk_apply_button.clicked.connect(self.rtk_apply)
        self.load_project_button.clicked.connect(self.loadProject)
        self.tr_dat_button.clicked.connect(self.getDatFile)
        self.common_apply_button.clicked.connect(self.apply_common_input)
        self.get_image_button.clicked.connect(self.get_image_folder)
        self.jpg_radio.toggled.connect(self.onRadioButtonToggled)
        self.jpeg_radio.toggled.connect(self.onRadioButtonToggled)
        self.png_radio.toggled.connect(self.onRadioButtonToggled)
        self.image_apply_button.clicked.connect(self.apply_image_settings)
        self.cif_button.clicked.connect(lambda: self.setLocation("cif"))
        self.shp_button.clicked.connect(lambda: self.setLocation("shp"))
        self.export_project_button.clicked.connect(self.saveProject)
        self.export_xlsx_button.clicked.connect(self.saveToExcel)
        self.table_widget.item_double_clicked.connect(self.on_item_double_clicked)
        self.image_editor.table_update_request.connect(self.on_table_update_request)
        self.update_code_button.clicked.connect(self.on_update_code)
        self.classify_image_button.clicked.connect(self.on_classify_image)
        self.change_mode_toggle.stateChanged.connect(self.change_mode)
        self.settings_button.clicked.connect(self.show_settings_dialog)

        for wgt in [self.band, self.rtk_table_widget, self.table_widget]:
            wgt.hide()

    
    @Slot(int, str, str, str)
    def on_item_double_clicked(self, row, num, path, name):
        orginal_image = os.path.join(path, name)
        self.status_message.setText(' '.join([num, orginal_image]))
        self.image_editor.open_image_from(orginal_image, set_current=True)
        self.image_editor.table_row = (row, num)
        self.image_editor.update_image()

    @Slot(int, str, str)
    def on_table_update_request(self, row, path, filename):
        self.table_widget.setCellItemAligned(row, self._headerindex("사진파일(경로)"), path)
        self.table_widget.setCellItemAligned(row, self._headerindex("사진파일명"), filename)

        ##############################################################
        # 위성사진 생성
        num = self.table_widget.item(row, 0).text()
        x = self.table_widget.item(row, 2).text()
        y = self.table_widget.item(row, 1).text()
        api_key =  self.settings.get_section("VWORLD_API_SETTINGS").get("api_key", '')
        if api_key == '':
            self.show_modal("error", parent=self.main_frame, title=" Empty API Key", description="vworld api key가 입력되지 않아 위성사진을 불러올 수 없습니다.")
            return

        vworld_dlg = VWorldMapViewer(float(x), float(y), name=num, path=path, apply_transform=True, row=row, parent=self, api_key=api_key)
        vworld_dlg.save_sat_image_request.connect(self.on_save_sat_image)
        vworld_dlg.exec()
        ##############################################################

        self.show_modal("success", parent=self.main_frame, title=" Successfully Apllied", description=f"성공적으로 적용/저장되었습니다.\n{filename}")

    @Slot(int, str)
    def on_save_sat_image(self, row, filename):
        self.table_widget.setCellItemAligned(row, self._headerindex("위성사진"), filename)

    @property
    def image_folder(self): 
        return self._image_folder

    @image_folder.setter
    def image_folder(self, folder):
        self._image_folder = folder

    def alignAllCellsCenter(self):
        for row in range(self.table_widget.rowCount()):
            for col in range(self.table_widget.columnCount()):
                item = QTableWidgetItem()
                item.setTextAlignment(Qt.AlignCenter)
                self.table_widget.setItem(row, col, item)

    # def add_menubar(self):
    #     menubar = self.menuBar()
    #     file_menu = menubar.addMenu('파일')

    #     new_action = QAction(QIcon(':resources/icons/document-add-svgrepo-com.svg'), '새 문서', self)
    #     # new_action.triggered.connect(self.new_document)
    #     file_menu.addAction(new_action)

    #     open_action = QAction(QIcon(':resources/icons/album-svgrepo-com.svg'), '열기', self)
    #     # open_action.triggered.connect(self.open_image)
    #     file_menu.addAction(open_action)

    #     save_action = QAction(QIcon(':resources/icons/diskette-svgrepo-com.svg'), '저장', self)
    #     # save_action.triggered.connect(self.save_image)
    #     file_menu.addAction(save_action)

    #     view_menu = menubar.addMenu('보기')
    #     self.show_layer_action = QAction(QIcon(':resources/icons/layers-svgrepo-com.svg'), '레이어', self)
    #     # self.show_layer_action.triggered.connect(self.show_layer_win)
    #     self.show_layer_action.setEnabled(False)
    #     self.show_explorer_action = QAction(QIcon(':resources/icons/library-svgrepo-com.svg'), "탐색기", self)
    #     # self.show_explorer_action.triggered.connect(self.show_explorer_win)
    #     self.show_explorer_action.setEnabled(False)
    #     self.show_preview_action = QAction(QIcon(':resources/icons/gallery-svgrepo-com.svg'), "미리보기", self)
    #     # self.show_preview_action.triggered.connect(self.show_preview_win)
    #     self.show_preview_action.setEnabled(False)
    #     view_menu.addAction(self.show_layer_action)
    #     view_menu.addAction(self.show_explorer_action)
    #     view_menu.addAction(self.show_preview_action)

    def setResizable(self, widget=None, width=100):
        if widget is None:
            widget = self
        widget.setMinimumSize(width, 150)
        widget.setMaximumSize(16777215, 16777215)

    def create_scroll_area(self, widget, parent=None):
        scroll_area = QScrollArea(parent)
        scroll_area.setWidgetResizable(True)  # 위젯 크기 조절 가능
        scroll_area.setWidget(widget)  # 주어진 위젯을 스크롤 영역에 설정
        return scroll_area

    def change_mode(self):
        self.table_widget.set_mode = self.change_mode_toggle.isChecked()
        self.table_widget.clearSelection()  
        is_maximized = self.isMaximized()      
        if self.change_mode_toggle.isChecked():
            self.mode = "edit-table"
            self.image_editor_frame.hide()
            self.sidemenu.show()
            self.table_widget.setSelectionBehavior(QAbstractItemView.SelectItems)
            self.table_widget.setSelectionMode(QAbstractItemView.ExtendedSelection)
            self.table_widget.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
            self.table_widget.show_all_columns()
            if is_maximized:
                table_size = sum([self.table_widget.columnWidth(col) for col in range(self.table_widget.columnCount())])  
                self.table_widget.setMinimumWidth(self.screen().size().width() - self.sidemenu.minimumWidth())  
                self.setFixedWidth(self.screen().size().width())                
                self.showMaximized()
            else:
                self.table_widget.setMinimumWidth(self.__width - self.sidemenu.minimumWidth())
                self.setMinimumWidth(self.__width)

            self.image_editor.initialize_pixmap()
        else:
            self.mode = "edit-image"
            self.sidemenu.hide()
            self.image_editor_frame.show()
            self.table_widget.setSelectionBehavior(QAbstractItemView.SelectRows)
            self.table_widget.setSelectionMode(QAbstractItemView.SingleSelection)
            self.table_widget.setEditTriggers(QAbstractItemView.NoEditTriggers) 
            self.table_widget.hide_columns(['X', 'Y', '도선등급', '도선명', '표지재질', '토지소재(동리)', '토지소재(지번)', '지적(임야)도', '설치년월일', '조사년월일', '조사자(직)', '조사자(성명)', '조사내용', '경위도(B)', '경위도(L)', '원점', '표고'])
            self.table_widget.setColumnWidth(self.table_widget.get_column_header().index("사진파일(경로)"), 350)
            table_size = sum([self.table_widget.columnWidth(col) for col in range(self.table_widget.columnCount()) if self.table_widget.isColumnHidden(col) == False])
            self.__width = self.size().width()
            self.table_widget.setFixedWidth(table_size)
        self.setResizable()

    def add_toolbar(self):
        self.toolbar = QToolBar()
        self.toolbar.setWindowTitle('MainToolbar')
        self.addToolBar(Qt.TopToolBarArea, self.toolbar)

        self.change_mode_toggle = CustomToggleButton()

        # 사용 설명서
        help_action = QAction(QIcon(':resources/icons/book-2.svg'), '사용설명서', self)
        help_action.triggered.connect(self.on_help)

        self.toolbar.addWidget(self.change_mode_toggle)
        self.toolbar.addAction(help_action)

    def add_dockableWidget(self, title:str, wdg:QWidget):
        dock = QDockWidget(title, self)
        temp_widget = QWidget(dock)
        dock.setTitleBarWidget(temp_widget)
        dock.setAllowedAreas(Qt.LeftDockWidgetArea)
        dock.setWidget(wdg)
        self.addDockWidget(Qt.LeftDockWidgetArea, dock)    
        return dock
    
    def on_help(self):
        # Help 파일 경로
        script_dir = os.path.dirname(os.path.abspath(__file__))  # 현재 스크립트 디렉토리
        help_file_path = os.path.join(script_dir, "help", "help.html")  # 절대 경로로 변환
        try:
            # 1. Chrome 브라우저에서 열기 시도
            chrome_paths = [r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe", 
                           r"C:\Program Files\Google\Chrome\Application\chrome.exe"]  # Chrome 실행 파일 경로
            chrome_path = None
            for _path in chrome_paths:
                if os.path.exists(_path):
                    chrome_path = _path
                    break
            webbrowser.register('chrome', None, webbrowser.BackgroundBrowser(chrome_path))
            webbrowser.get('chrome').open(help_file_path)
        except Exception as chrome_error:
            print(f"Failed to open in Chrome: {chrome_error}")

            try:
                # 2. 기본 브라우저에서 열기 시도
                webbrowser.open(help_file_path)
                print("Opened in default browser")
            except Exception as default_error:
                print(f"Failed to open in default browser: {default_error}")
                print("Error opening help file.")
                self.show_modal("error", parent=self.main_frame, title=" Cannot Open Help", description=f"파일 열기 중 오류 발생:\n{default_error}")

    def show_modal(self, modal_type, **kargs):
        """ 메시지 송출 """
        default_settings = {'position': 'bottom-right', 'duration': 2000, 'closeIcon': ':resources/icons/x.svg'}
        modal_collection = {'info':QCustomModals.InformationModal, 
                            'success':QCustomModals.SuccessModal, 
                            'error':QCustomModals.ErrorModal, 
                            'warning':QCustomModals.WarningModal, 
                            'custom':QCustomModals.CustomModal}
        if modal_type not in ['info', 'success', 'error', 'warning', 'custom']: modal_type = 'custom'

        for key, value in kargs.items():
            default_settings[key] = value    

        modal = modal_collection[modal_type](**default_settings)
        modal.show()

    def input_rtkdata_toggle(self):
        widgets = [self.rtk_data_sub, self.rtk_table_widget, self.save_table, self.table_to_tr, self.column_hide, self.fill_number]
        self.band.show()
        if self.input_rtkdata_button.isChecked():
            self.table_widget.hide()           
            for widget in widgets:
                widget.show()
        else:
            self.table_widget.show()
            for widget in widgets: 
                widget.hide()

        self.input_data_sub.show() if self.input_data_button.isChecked() else self.input_data_sub.hide()
        self.notice_frame.hide()

    def rtk_table_hide_column(self):
        for col in [4, 5, 8, 12, 14, 15, 16, 18, 19]:
            if self.column_hide.isChecked():     
                self.rtk_table_widget.setColumnHidden(col, True)
                self.column_hide.setToolTip('컬럼 보이기')
            else:
                self.rtk_table_widget.setColumnHidden(col, False)
                self.column_hide.setToolTip('컬럼 숨기기')

    def auto_fill_number(self):
        if self.rtk_data_file is None:
            self.show_modal("error", parent=self.main_frame, title=" NOT Input RTK Data", description="RTK 데이터가 입력되지 않았습니다.")
            return
        
        start_number, valid = QInputDialog.getInt(self, "자동번호 입력", "시작 번호", 1, 1, 999999, 1)
        if valid:
            for row in range(0, self.rtk_table_widget.rowCount(), 2):
                self.rtk_table_widget.item(row, 0).setText(str(start_number))
                start_number += 1


    # 리소스에서 파일을 복사하는 함수
    def copy_resource_to_file(self, resource_path, destination_path):
        resource = QFile(resource_path)
        if not resource.open(QIODevice.OpenModeFlag.ReadOnly):
            print(f"리소스 파일을 열 수 없습니다: {resource_path}")
            return  False

        with open(destination_path, 'wb') as dest_file:
            dest_file.write(resource.readAll())
        
        return True

    def loadRTKdata(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Excel 파일 선택", "", "Excel Files (*.xlsx)")
        self.rtk_data_file = file_path
        self.rtk_data_path = os.path.dirname(file_path)

        if file_path:
            try:
                # Load the workbook and the first sheet using openpyxl
                wb = load_workbook(file_path)
                sheet = wb.active  # Get the active sheet

                # Extract data from the sheet into a list of lists
                data = []
                for row in sheet.iter_rows(values_only=True):  # Only get values, not cell objects
                    data.append(list(row))  # Convert row tuple to list

                if not data:
                    raise ValueError("The Excel file is empty or has no readable data.")

                # Convert the data to a Pandas DataFrame
                headers = data[0]  # First row as headers
                rows = data[1:]  # Remaining rows as data
                if not headers:
                    raise ValueError("The Excel file does not contain headers in the first row.")
                df = pd.DataFrame(rows, columns=headers)

                # Sort DataFrame by '번호' column
                if '번호' not in df.columns:                 
                    raise ValueError("'번호' column not found in the Excel file.")

                df_sorted = df.sort_values(by='번호')

                # Prepare data for TableWidget (include headers)
                data_for_table = [df_sorted.columns.tolist()] + df_sorted.values.tolist()

                # Clear existing content
                self.rtk_table_widget.clearContents()
                self.rtk_table_widget.setRowCount(len(data_for_table) - 1)
                self.rtk_table_widget.alignAllCellsCenter()

                # Populate table
                for row_idx, row in enumerate(data_for_table[1:], start=0):
                    for col_idx, value in enumerate(row):
                        item = QTableWidgetItem(str(value) if value is not None else "")
                        item.setTextAlignment(Qt.AlignCenter)
                        self.rtk_table_widget.setItem(row_idx, col_idx, item)

            except ValueError as e:
                self.show_modal("error", parent=self.main_frame, title=" Invalid Excel File", description=f"Please enter a valid excel file(rtk) to proceed.\n{e}")
                self.status_message.setText(f"파일 로드 중 오류 발생: {e}")
                return
            
    def table_to_trdat(self):
        """ tr.dat 파일 생성 """
        if self.rtk_data_path is not None:
            filename = os.path.join(self.rtk_data_path, 'temp_tr.dat')
        else:
            self.show_modal("error", parent=self.main_frame, title=" NOT Input RTK Data", description="Please enter an excel file(rtk) to proceed.")
            self.status_message.setText("RTK data가 입력되지 않았습니다.")
            return
        
        # Copy the source sheet data to the new workbook
        data = []
        for row in range(0, self.rtk_table_widget.rowCount(), 2):
            row_items = []
            for col in [0, 9, 10]:     # 번호, x, y
                item = self.rtk_table_widget.item(row, col)
                row_items.append(item.text())
            data.append(row_items)
            
        df = pd.DataFrame(data, columns=['번호', 'X', 'Y'])
        df['X'] = df['X'].apply(lambda x: int(np.round(float(x), 2) * 100))
        df['Y'] = df['Y'].apply(lambda x: int(np.round(float(x), 2) * 100))

        with open(filename, 'w', encoding='euc-kr') as file:
            for i in range(len(df)):
                row_data = df.iloc[i].values.tolist()
                file.write(f'{row_data[0]}\t{row_data[1]}\t{row_data[2]}\n')

        self.show_modal("success", parent=self.main_frame, title=" Saving Success", description=f"tr.dat 파일이 성공적으로 저장되었습니다.\n{filename}")
        self.status_message.setText(f"tr.dat 파일이 성공적으로 저장되었습니다: {filename}")

    def save_rtk_table(self):
        """ RTK 데이터 재생성 """
        if self.rtk_data_file is None:
            self.show_modal("error", parent=self.main_frame, title=" NOT Input RTK Data", description="RTK 데이터가 입력되지 않았습니다.")
            return
        
        wb = load_workbook(self.rtk_data_file)
        sheet = wb.active
        sheet.delete_rows(2, sheet.max_row)

        # Copy the source sheet data to the new workbook
        for row in range(0, self.rtk_table_widget.rowCount()):
            row_items = []
            for col in range(self.rtk_table_widget.columnCount()-4):
                if row%2==1 and col==0:
                    item = self.rtk_table_widget.item(row-1, col)
                    item_value = f'{item.text()}(재관측)'
                else:
                    item = self.rtk_table_widget.item(row, col)
                    item_value = item.text()
                row_items.append(item_value)
            for idx, item in enumerate(row_items):
                sheet.cell(row=row+2, column=idx+1, value=item)
       
        # Save the new workbook
        time_stamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        savas_filename = os.path.join(self.rtk_data_path, f'{time_stamp}_data.xlsx')
        wb.save(savas_filename)
        wb.close()
        self.show_modal("success", parent=self.main_frame, title=" Saving Success", description=f"RTK_data가 성공적으로 생성되었습니다.\n{savas_filename}")
        self.status_message.setText(f"Save RTK_data_table to '{savas_filename}' successfully.")
    
    def rtk_sort(self):
        self.rtk_table_widget.sort_table_widget(0,1)

    def rtk_timecheck(self):
        # 단일 관측시간 점검
        survey_count = self.rtk_table_widget.rowCount()

        for row in range(survey_count):
            t1 = self.rtk_table_widget.item(row, 1).text().strip()
            t2 = self.rtk_table_widget.item(row, 2).text().strip()
            t_diff = self.calculate_time_difference(t1, t2, unit='seconds')

            if t_diff < 60:
                self.rtk_table_widget.item(row, 2).setText(self.add_time_to_datetime(t1, seconds= random.randrange(61, 70)))
                self.rtk_table_widget.item(row, 2).setForeground(Qt.red)
        # 관측간 간격
        for row in range(0, survey_count, 2):
            t1 = self.rtk_table_widget.item(row, 1).text().strip()
            t2 = self.rtk_table_widget.item(row+1, 1).text().strip()
            t_diff = self.calculate_time_difference(t1, t2, unit='seconds')
            if t_diff < 3600:
                new_time = self.add_time_to_datetime(t1, seconds= random.randrange(3661, 3900))
                self.rtk_table_widget.item(row+1, 1).setText(new_time)
                self.rtk_table_widget.item(row+1, 2).setText(self.add_time_to_datetime(new_time, seconds = random.randrange(61,70)))
                self.rtk_table_widget.item(row+1, 1).setForeground(Qt.red)
                self.rtk_table_widget.item(row+1, 2).setForeground(Qt.red)

    def add_time_to_datetime(self, datetime_str, hours=0, minutes=0, seconds=0):
        # 문자열을 datetime 객체로 변환
        original_datetime = datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S")
        
        # timedelta를 사용하여 시간과 분을 더함
        new_datetime = original_datetime + timedelta(hours=hours, minutes=minutes, seconds=seconds)
        
        # 새로운 datetime 객체를 문자열로 변환하여 반환
        return new_datetime.strftime("%Y-%m-%d %H:%M:%S")

    def calculate_time_difference(self, datetime1, datetime2, unit='seconds'):
        """
        두 datetime 문자열의 차이를 계산합니다.

        :param datetime1: 첫 번째 datetime 문자열 (예: "2024-11-20 14:20:02")
        :param datetime2: 두 번째 datetime 문자열 (예: "2024-11-20 15:20:02")
        :param unit: 결과를 반환할 단위 ('seconds', 'minutes', 'hours')
        :return: 시간 차이 (int 또는 float)
        """
        # datetime 문자열을 datetime 객체로 변환
        format_str = "%Y-%m-%d %H:%M:%S"
        dt1 = datetime.strptime(datetime1, format_str)
        dt2 = datetime.strptime(datetime2, format_str)
        
        # 시간 차이 계산
        delta = dt2 - dt1  # 시간 차이 계산
        
        # 단위에 따라 반환
        if unit == 'seconds':
            return delta.total_seconds()
        elif unit == 'minutes':
            return delta.total_seconds() / 60
        elif unit == 'hours':
            return delta.total_seconds() / 3600
        else:
            raise ValueError("Invalid unit. Choose 'seconds', 'minutes', or 'hours'.")

    def rtk_location(self):
        jijuk, _ = QFileDialog.getOpenFileName(self,"Get Jijuk DB", "", "지적파일 (*.cif; *.shp)")
        try:
            if jijuk:
                if jijuk.lower().endswith(".cif"):
                    gdf = CifGeoDataFrame(jijuk).convert_to_geodataframe()
                elif jijuk.lower().endswith(".shp"):
                    gdf = gpd.read_file(jijuk, encoding='euc-kr')
          
                if gdf is None: 
                    return
                
                not_found = []
                shp_from = "landygo" if "GKEY" in gdf.columns else "general"  # GKEY가 있으면 랜디고에서, 아니면 일반 LXGIS
                
                for i in range(self.rtk_table_widget.rowCount()):
                    if shp_from == "landygo":
                        location = find_attributes_containing_point(gdf, (float(self.rtk_table_widget.item(i, 10).text()), 
                                                                          float(self.rtk_table_widget.item(i, 9).text())), ["PNU", "GKEY"]) 
                    else:
                        location = find_attributes_containing_point(gdf, (float(self.rtk_table_widget.item(i, 10).text()), 
                                                                          float(self.rtk_table_widget.item(i, 9).text())), ["PNU", "DOHO"])

                    if location is not None:
                        pnu, dom = location.iloc[0, :]
                        self.rtk_table_widget.item(i, 21).setText(CifGeoDataFrame().getDistrictName(pnu))
                        self.rtk_table_widget.item(i, 22).setText(CifGeoDataFrame().pnu2jibun(pnu))
                        if shp_from == "landygo":
                            self.rtk_table_widget.item(i, 23).setText(f'{str(dom)[-2:]}') 
                        else:
                            self.rtk_table_widget.item(i, 23).setText(self.dom_to_doho(dom)) 
                    else:
                        not_found.append(self.rtk_table_widget.item(i, 0).text())
                        self.rtk_table_widget.item(i, 21).setText("")
                        self.rtk_table_widget.item(i, 22).setText("")
                        self.rtk_table_widget.item(i, 23).setText("")

                message = "주소검색을 마쳤습니다."
                if not_found:
                    message = f"주소검색을 마쳤습니다.\n미작성된 소재지{str(not_found)}를 확인하세요."
                
                self.show_modal("success", parent=self.main_frame, title=" Searching is Done", description=message)
                self.status_message.setText(f"{message}")

        except Exception as e:
            self.show_modal("error", parent=self.main_frame, title=" Unknown Error", description=f"{e}")
            self.status_message.setText(str(e))

    def rtk_cover(self):
        """ 표지 작성 """
        record_file = '_표지.xlsx'
        template_path = self.RTK_TEMPLATE
        sheet_name = '@표지'

        try:
            # Load the template workbook
            copy_success = self.copy_resource_to_file(template_path, record_file)
            if not copy_success:
                self.status_message.setText("[표지] 파일 복사에 실패했습니다.")
                raise FileExistsError("[표지] 파일 복사에 실패했습니다.")
        except FileExistsError:
            self.show_modal("error", parent=self.main_frame, title=" File Copy Failed", description=f"[표지] 파일 복사에 실패했습니다.")
            return
            
        new_wb = load_workbook(record_file)
        
        # Get the sheet to copy
        record_sheet = new_wb[sheet_name]

        # Remove the default sheet created in the new workbook
        for sht in new_wb.sheetnames:
            if sheet_name != sht:
                del new_wb[sht]        

        survey_count = self.rtk_table_widget.rowCount() // 2
        reception_number = self.reception_number.text().strip()
        ref_jibun = self.ref_jibun.text().strip()
        jiguname = self.jigu_name.text().strip()

        if not reception_number.endswith('호'):
            reception_number = reception_number + '호'        
        if not ref_jibun.endswith('번지'):
            ref_jibun = ref_jibun + '번지'
        if not jiguname.endswith('지구'):
            jiguname = ""
        else:
            jiguname = f'[ {jiguname} ]'

        # 관측 정보 입력
        record_sheet['D15'].value = ref_jibun                       # 대표 지번
        record_sheet['E17'].value = jiguname                        # 지구명
        record_sheet['N11'].value = self.ccp_type.currentText()     # 구분
        record_sheet['I19'].value = format_date_to_korean(self.rtk_table_widget.item(0,1).text().split(' ')[0])  # 관측일자
        record_sheet['N12'].value = f'{survey_count}점'             # 점 수
        record_sheet['I21'].value = f'{self.survey_license.currentText()}  {self.surveyor_name.currentText()}'  # 측량자
        record_sheet['N10'].value = reception_number                # 접수번호
        
        #print setting
        record_sheet.page_setup.orientation = record_sheet.ORIENTATION_PORTRAIT  # 가로 방향 설정
        record_sheet.page_setup.paperSize = record_sheet.PAPERSIZE_A4  # A4 용지 크기 설정
        record_sheet.print_options.horizontalCentered = True  # 가로 중앙 정렬 설정
                    
        # Save the new workbook
        time_stamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        savas_filename = os.path.join(self.rtk_data_path, f'{time_stamp}_표지.xlsx')
        new_wb.save(savas_filename)
        new_wb.close()
        self.show_modal("success", parent=self.main_frame, title=" Saving Success", description=f"[표지]를 성공적으로 저장하였습니다.\n{savas_filename}")
        self.status_message.setText(f"Save to '{savas_filename}' successfully.")

    def rtk_record(self):
        """ 위성관측기록부 작성 """
        record_file = '_관측기록부.xlsx'
        template_path = self.RTK_TEMPLATE
        sheet_name = '@관측기록부'

        try:
            # Load the template workbook
            copy_success = self.copy_resource_to_file(template_path, record_file)
            if not copy_success:
                self.status_message.setText("[관측기록부] 파일 복사에 실패했습니다.")
                raise FileExistsError("[관측기록부] 파일 복사에 실패했습니다.")
        except FileExistsError:
            self.show_modal("error", parent=self.main_frame, title=" File Copy Failed", description=f"[관측기록부] 파일 복사에 실패했습니다.")
            return
            
        new_wb = load_workbook(record_file)
        
        # Get the sheet to copy
        record_sheet = new_wb[sheet_name]

        # Remove the default sheet created in the new workbook
        for sht in new_wb.sheetnames:
            if sheet_name != sht:
                del new_wb[sht]        

        survey_count = self.rtk_table_widget.rowCount()

        copy_row_with_merge(record_sheet, 17, 17, survey_count-1)

        try:
            if survey_count % 2 == 1:
                self.status_message.setText('관측기록이 짝수 회가 아닙니다.')
                raise ValueError('관측기록이 짝수 회가 아닙니다.')
        except ValueError:
            self.show_modal("error", parent=self.main_frame, title=" Invalid Survey Count", description=f"관측기록이 짝수 회가 아닙니다.\n검사 후 다시 시도하세요.")
            return

        # set font, alignment, border style
        set_font(record_sheet[f"B17:P{17 + survey_count-1}"], sz=9)
        set_alignment(record_sheet[f"B17:P{17 + survey_count-1}"], horizontal='center', vertical='center')
        set_border(record_sheet[f"B17:P{17 + survey_count-1}"], edges=["all"], border_style='thin')
        
        for col in range(17, 17+ survey_count):
            record_sheet.row_dimensions[col].height = 15

        # 관측 정보 입력
        record_sheet['E3'].value = self.jigu_name.text()  # 지구명
        record_sheet['E4'].value = format_date_to_korean(self.rtk_table_widget.item(0,1).text().strip().split(' ')[0])  # 관측일자
        record_sheet['N4'].value = self.surveyor_name.currentText()  # 관측자
        record_sheet['E5'].value = self.jigu_attr.text()  # 지구특성
        record_sheet['N9'].value = self.rtk_table_widget.item(0,16).text()  # 수신기명
        record_sheet['N10'].value = self.antena_sn.text()  # 안테나 명(번호)
        
        # Copy the source sheet data to the new workbook
        for row in range(survey_count):
            row_items = []
            for col in range(self.rtk_table_widget.columnCount()):
                item = self.rtk_table_widget.item(row, col)
                row_items.append(item.text().strip())

            record_sheet[f'B{17+row}'].value = row_items[0]         # 번호
            record_sheet[f'C{17+row}'].value = row % 2 + 1          # 세션
            record_sheet[f'D{17+row}'].value = row_items[1][2:]     # 관측시간-시작
            record_sheet[f'G{17+row}'].value = row_items[2][2:]     # 관측시간-종료
            record_sheet[f'I{17+row}'].value = row_items[4]         # 수평
            record_sheet[f'K{17+row}'].value = row_items[5]         # 수직
            record_sheet[f'L{17+row}'].value = 1.8                  # 안테나고
            record_sheet[f'M{17+row}'].value = f'{float(row_items[13]):.2f}'  # PDOP
            record_sheet[f'O{17+row}'].value = row_items[17]        # 위성수
            record_sheet[f'P{17+row}'].value = '15˚'                # 위성고도각

        max_row = record_sheet.max_row
        rng = record_sheet[f'A17:P{max_row}']
        set_font(rng, sz=8, name="굴림")

        # cell merge
        for row in range(17, 17+ survey_count, 2):
            record_sheet.merge_cells(f'B{row}:B{row+1}')
        
        #print setting
        record_sheet.page_setup.orientation = record_sheet.ORIENTATION_PORTRAIT  # 가로 방향 설정
        record_sheet.page_setup.paperSize = record_sheet.PAPERSIZE_A4  # A4 용지 크기 설정
        record_sheet.page_margins = PageMargins(left=0.2, right=0.2, top=0.2, bottom=0.2, header=0, footer=0)  # 페이지 여백 설정
        record_sheet.print_options.horizontalCentered = True  # 가로 중앙 정렬 설정
        record_sheet.print_area = None  # 인쇄 영역 설정
                    
        # Save the new workbook
        time_stamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        savas_filename = os.path.join(self.rtk_data_path, f'{time_stamp}_관측기록부.xlsx')
        new_wb.save(savas_filename)
        new_wb.close()
        self.show_modal("success", parent=self.main_frame, title=" Saving Success", description=f"[관측기록부]가 성공적으로 저장되었습니다.\n{savas_filename}")
        self.status_message.setText(f"'{sheet_name}' sheet copied to '{savas_filename}' successfully.")

    def rtk_result(self):
        """ 위성관측결과부 작성 """
        record_file = '_관측결과부.xlsx'
        template_path = self.RTK_TEMPLATE
        sheet_name = '@관측결과부'

        try:
            # Load the template workbook
            copy_success = self.copy_resource_to_file(template_path, record_file)
            if not copy_success:
                self.status_message.setText("[관측결과부] 파일 복사에 실패했습니다.")
                raise FileExistsError("[관측결과부] 파일 복사에 실패했습니다.")
            
        except FileExistsError:
            self.show_modal("error", parent=self.main_frame, title=" File Copy Failed", description=f"[관측결과부] 파일 복사에 실패했습니다.")
            return  
        
        new_wb = load_workbook(record_file)
        
        # Get the sheet to copy
        record_sheet = new_wb[sheet_name]

        # Remove the default sheet created in the new workbook
        for sht in new_wb.sheetnames:
            if sheet_name != sht:
                del new_wb[sht]        

        survey_count = self.rtk_table_widget.rowCount()//2
        copy_row_with_merge(record_sheet, 17, 17, survey_count-1)

        # set font, alignment, border style
        set_font(record_sheet[f"B17:P{17 + survey_count-1}"], sz=9)
        set_alignment(record_sheet[f"B17:P{17 + survey_count-1}"], horizontal='center', vertical='center')
        set_border(record_sheet[f"B17:P{17 + survey_count-1}"], edges=["all"], border_style='thin')
        
        for col in range(17, 17+ survey_count):
            record_sheet.row_dimensions[col].height = 20

        # 관측 정보 입력
        record_sheet['E3'].value = "세계측지계"  # 지구명
        record_sheet['E4'].value = self.jigu_name.text()  # 지구명
        record_sheet['E5'].value = format_date_to_korean(self.rtk_table_widget.item(0,1).text().strip().split(' ')[0])  # 관측일자
        record_sheet['M5'].value = self.surveyor_name.currentText()  # 관측자
        record_sheet['E6'].value = "중부원점"  # 투영원점
        
        # Copy the source sheet data to the new workbook
        for row in range(0, self.rtk_table_widget.rowCount(), 2):
            row_items = []
            for col in range(self.rtk_table_widget.columnCount()):
                item = self.rtk_table_widget.item(row, col)
                row_items.append(item.text())

            record_sheet[f'B{17+row//2}'].value = row // 2 + 1   # 순번
            record_sheet[f'C{17+row//2}'].value = row_items[0]   # 점명
            record_sheet[f'D{17+row//2}'].value = convert_decimal_to_roundup_angle(row_items[6])   # 위도
            record_sheet[f'F{17+row//2}'].value = convert_decimal_to_roundup_angle(row_items[7])   # 경도
            record_sheet[f'H{17+row//2}'].value = row_items[8]   # 타원체고
            record_sheet[f'J{17+row//2}'].value = f'{np.round(float(row_items[9]), 2):.2f}'   # X
            record_sheet[f'L{17+row//2}'].value = f'{np.round(float(row_items[10]), 2):.2f}'  # Y
            record_sheet[f'N{17+row//2}'].value = row_items[11]  # 표고
            record_sheet[f'P{17+row//2}'].value = ''             # 비고 

        max_row = record_sheet.max_row
        rng = record_sheet[f'A17:P{max_row}']
        set_font(rng, sz=8, name="굴림")
        
        #print setting
        record_sheet.page_setup.orientation = record_sheet.ORIENTATION_PORTRAIT  # 가로 방향 설정
        record_sheet.page_setup.paperSize = record_sheet.PAPERSIZE_A4  # A4 용지 크기 설정
        record_sheet.page_margins = PageMargins(left=0.2, right=0.2, top=0.2, bottom=0.2, header=0, footer=0)  # 페이지 여백 설정
        record_sheet.print_options.horizontalCentered = True  # 가로 중앙 정렬 설정
        record_sheet.print_area = None  # 인쇄 영역 설정

        # Save the new workbook
        time_stamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        savas_filename = os.path.join(self.rtk_data_path, f'{time_stamp}_관측결과부.xlsx')
        new_wb.save(savas_filename)
        new_wb.close()
        self.show_modal("success", parent=self.main_frame, title=" Saving Success", description=f"[관측결과부]가 성공적으로 저장되었습니다.\n{savas_filename}")
        self.status_message.setText(f"'{sheet_name}' sheet copied to '{savas_filename}' successfully.")

    def rtk_ilram(self):
        """ 지적기준점 일람표 작성 """
        record_file = '_기준점일람표.xlsx'
        template_path = self.RTK_TEMPLATE
        sheet_name = '@기준점일람표'

        try:
            # Load the template workbook
            copy_success = self.copy_resource_to_file(template_path, record_file)
            if not copy_success:
                self.status_message.setText("[기준점일람표] 파일 복사에 실패했습니다.")
                raise FileExistsError("[기준점일람표] 파일 복사에 실패했습니다.")
        except FileExistsError:
            self.show_modal("error", parent=self.main_frame, title=" File Copy Failed", description=f"[기준점일람표] 파일 복사에 실패했습니다.")
            return
        
        new_wb = load_workbook(record_file)
        
        # Get the sheet to copy
        record_sheet = new_wb[sheet_name]

        # Remove the default sheet created in the new workbook
        for sht in new_wb.sheetnames:
            if sheet_name != sht:
                del new_wb[sht]        

        survey_count = self.rtk_table_widget.rowCount()//2
        copy_row_with_merge(record_sheet, 17, 17, survey_count-1)

        # set font, alignment, border style
        set_font(record_sheet[f"A4:O{4 + survey_count-1}"], sz=9)
        set_alignment(record_sheet[f"A4:O{4 + survey_count-1}"], horizontal='center', vertical='center')
        set_border(record_sheet[f"A4:O{4 + survey_count-1}"], edges=["all"], border_style='thin')

        # Copy the source sheet data to the new workbook
        for row in range(0, self.rtk_table_widget.rowCount(), 2):
            row_items = []
            for col in range(self.rtk_table_widget.columnCount()):
                item = self.rtk_table_widget.item(row, col)
                row_items.append(item.text())

            record_sheet[f'A{4+row//2}'].value = row // 2 + 1   # 연번
            record_sheet[f'B{4+row//2}'].value = "지적도근점"   # 명칭
            record_sheet[f'C{4+row//2}'].value = row_items[0]   # 번호(점명)
            record_sheet[f'D{4+row//2}'].value = "신설"   # 구분 
            record_sheet[f'E{4+row//2}'].value = f'{np.round(float(row_items[9]), 2):.2f}'   # X
            record_sheet[f'F{4+row//2}'].value = f'{np.round(float(row_items[10]), 2):.2f}'  # Y
            record_sheet[f'G{4+row//2}'].value = "중부"   # 원점명 
            record_sheet[f'H{4+row//2}'].value = row_items[11]  # 표고
            record_sheet[f'I{4+row//2}'].value = convert_decimal_to_roundup_angle(row_items[6])   # 위도
            record_sheet[f'J{4+row//2}'].value = convert_decimal_to_roundup_angle(row_items[7])   # 경도
            record_sheet[f'K{4+row//2}'].value = ' '.join([row_items[21], row_items[22]]).replace('경기도 용인시 처인구 ', '')   # 소재지
            record_sheet[f'L{4+row//2}'].value = row_items[1].strip().split(' ')[0]   # 설치일자
            record_sheet[f'M{4+row//2}'].value = row_items[20]  # 재질
            record_sheet[f'N{4+row//2}'].value = ''   # 비고
            record_sheet[f'O{4+row//2}'].value = self.surveyor_name.currentText()  # 팀명

        set_border(record_sheet[f"A1:O{4 + survey_count-1}"], edges=["outer"], border_style='medium', reset=False)

        #print setting
        record_sheet.page_setup.orientation = record_sheet.ORIENTATION_LANDSCAPE  # 가로 방향 설정
        record_sheet.page_setup.paperSize = record_sheet.PAPERSIZE_A4  # A4 용지 크기 설정
        record_sheet.page_setup.fitToHeight = 0  # 한 페이지에 모든 열 맞춤 X
        record_sheet.page_setup.fitToWidth = 1  # 한 페이지에 모든 행 맞춤

        # Save the new workbook
        time_stamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        savas_filename = os.path.join(self.rtk_data_path, f'{time_stamp}_기준점일람표.xlsx')
        new_wb.save(savas_filename)
        new_wb.close()
        self.show_modal("success", parent=self.main_frame, title=" Saving Success", description=f"[기준점일람표]가 성공적으로 저장되었습니다.\n{savas_filename}")
        self.status_message.setText(f"'{sheet_name}' sheet copied to '{savas_filename}' successfully.")
    
    def rtk_report_all(self):
        self.rtk_cover()
        self.rtk_record()
        self.rtk_result()
        self.rtk_ilram()

    def rtk_apply(self):
        # 공통값 입력 
        if self.rtk_data_file is None:
            self.show_modal("error", parent=self.main_frame, title=" Error", description="먼저 RTK 데이터 파일을 입력해주세요.")
            return
        
        self.install_date_input.setText(self.rtk_table_widget.item(self.rtk_table_widget.rowCount()-1, 1).text().strip().split(' ')[0])
        self.survey_date_input.setText(self.rtk_table_widget.item(self.rtk_table_widget.rowCount()-1, 1).text().strip().split(' ')[0])
        self.surveyor_position_input.setText(self.surveyor_grade.currentText())
        self.surveyor_input.setText(self.surveyor_name.currentText())
        self.findings_input.setText('신설')
        self.origin_input.setPlaceholderText('원점')
        self.origin_input.setText('세계측지계(중부)')

        # copy rtk_table to another table
        self.table_widget.clearContents()
        row_count = self.rtk_table_widget.rowCount()//2
        self.table_widget.setRowCount(row_count)

        headers = {'번호':'점번호', 'X':'X', 'Y':'Y', '위도':'경위도(B)', '경도':'경위도(L)', 'Z':'표고', '토지소재(동리)':'토지소재(동리)', 
                   '토지소재(지번)':'토지소재(지번)', '지적(임야)도':'지적(임야)도', '재질':'표지재질'}
        
        for rtk_header, table_header in headers.items():
            self.copy_columns(self.rtk_table_widget, self.table_widget, rtk_header, table_header)

    def copy_columns(self, source_table: QTableWidget, target_table: QTableWidget, source_header: str, target_header: str = None):
        """
        source_table의 "Column A" 데이터를 target_table의 "Column A"에 복사합니다.
        
        :param source_table: 복사할 데이터를 가지고 있는 QTableWidget
        :param target_table: 데이터를 붙여넣을 QTableWidget
        :param column_a_index: "Column A"의 인덱스
        """
        target_header = source_header if target_header is None else target_header
        row_count = self.rtk_table_widget.rowCount()

        # source_table의 "Column A" 데이터를 복사하여 target_table에 붙여넣기
        #rtk_header index
        rtk_index = self._headerindex(source_header, self.RTK_HEADERS)
        #table_header index
        table_index = self._headerindex(target_header, self.HEADER_LABELS)
        for row in range(0, row_count, 2):
            item = source_table.item(row, rtk_index)
            if item:  # 값이 있을 경우
                new_item = QTableWidgetItem(item)
            else:  # 값이 없을 경우 빈 셀로 처리
                new_item = QTableWidgetItem("")
                new_item.setTextAlignment(Qt.AlignCenter)
            
            target_table.setItem(row//2, table_index, new_item)
        
        self.common_input_button.setChecked(True)

    def load_table_from_pickle(self, table_widget: QTableWidget, file_path: str):
        try:
            # pickle 파일에서 데이터 로드
            with open(file_path, mode='rb') as file:
                data = pickle.load(file)

            # QTableWidget 초기화
            table_widget.setRowCount(0)  # 기존 데이터 초기화
            table_widget.setColumnCount(0)

            # 데이터가 비어있지 않은 경우
            if data:
                # 열 수 설정
                column_count = len(data[0])
                table_widget.setColumnCount(column_count)

                # 헤더 설정
                headers = list(data[0].keys())
                table_widget.setHorizontalHeaderLabels(headers)

                # 데이터 추가
                for row_data in data:
                    row_index = table_widget.rowCount()
                    table_widget.insertRow(row_index)
                    for column_index, header in enumerate(headers):
                        item = QTableWidgetItem(row_data[header] if row_data[header] is not None else "")
                        table_widget.setItem(row_index, column_index, item)
                        item.setTextAlignment(Qt.AlignCenter)
            
            self.status_message.setText(f"파일이 성공적으로 로드되었습니다: {file_path}")

        except Exception as e:
            self.show_modal("error", parent=self.main_frame, title=" Cannot Load Pickle", description=f"파일 로드 중 오류 발생:\n{e}")
            self.status_message.setText(f"파일 로드 중 오류 발생: {e}")    

    def loadProject(self):
        options = QFileDialog.Options()
        fileName, _ = QFileDialog.getOpenFileName(self, "Load project", "", "Pickle Files (*.pickle)", options=options)
        if fileName:
            self.load_table_from_pickle(self.table_widget, fileName)
        self.table_widget.setColumnWidth(self.table_widget.get_column_header().index("토지소재(동리)"), 200)
        self.table_widget.setColumnWidth(self.table_widget.get_column_header().index("사진파일(경로)"), 350)

    def save_table_to_pickle(self, table_widget: QTableWidget, file_path: str):
        try:
            data = []
            row_count = table_widget.rowCount()
            column_count = table_widget.columnCount()

            # 헤더 가져오기
            headers = []
            for column in range(column_count):
                headers.append(table_widget.horizontalHeaderItem(column).text())

            # 데이터 가져오기
            for row in range(row_count):
                row_data = {}
                for column in range(column_count):
                    item = table_widget.item(row, column)
                    if item is not None:
                        row_data[headers[column]] = item.text()
                    else:
                        row_data[headers[column]] = None  # 빈 셀은 None으로 설정
                data.append(row_data)

            # pickle 파일로 저장
            with open(file_path, mode='wb') as file:
                pickle.dump(data, file)
            self.show_modal("success", parent=self.main_frame, title=" Saving Success", description=f"테이블이 성공적으로 저장되었습니다.\n{file_path}")
            self.status_message.setText(f"파일이 성공적으로 저장되었습니다: {file_path}")

        except Exception as e:
            self.show_modal("error", parent=self.main_frame, title=" Cannot Save Pickle", description=f"파일 저장 중 오류 발생:\n{e}")
            self.status_message.setText(f"파일 저장 중 오류 발생: {e}") 
    
    def getDatFile(self):
        options = QFileDialog.Options()
        fileName, _ = QFileDialog.getOpenFileName(self, "Open TR.DAT File", "", "DAT Files (*.dat);;All Files (*)", options=options)
        if fileName:
            try:
                self.tr = fileName
                self.loadDataToTable(fileName)
                self.show_modal("success", parent=self.main_frame, title=" Loading Success", description=f"tr.dat 파일이 성공적으로 로드되었습니다.\n{fileName}")
                self.status_message.setText(f"파일이 성공적으로 로드되었습니다: {fileName}")
            except Exception as e:
                self.show_modal("error", parent=self.main_frame, title=" Cannot Load tr.dat", description=f"파일 로드 중 오류 발생:\n{e}")

    def loadDataToTable(self, fileName):
        with open(fileName, 'r') as file:
            lines = file.readlines()            
        self.rowCount = len(lines)
        self.table_widget.setRowCount(self.rowCount)
        
        for row, line in enumerate(lines):
            # line의 양끝 공백문자를 모두 제거한 후 tab을 모두 공백으로 변경
            line = line.replace('\t', ' ')
            # line에 있는 중복된 공백을 공백하나로 변경
            while True:
                if "  " in line:
                    line = line.replace("  ",' ')
                else:
                    break
            data = line.strip().split(' ')
            if len(data) >= 3:
                self.table_widget.setCellItemAligned(row, 0, data[0])  # 점번호
                self.table_widget.setCellItemAligned(row, 1, f"{float(data[1])/100:.2f}")  # X
                self.table_widget.setCellItemAligned(row, 2, f"{float(data[2])/100:.2f}")  # Y
                # 나머지는 빈 문자
                for i in range(3, self.table_widget.columnCount()+1):
                    self.table_widget.setCellItemAligned(row, i, '')
        self.table_widget.setColumnWidth(self.table_widget.get_column_header().index("토지소재(동리)"), 200)
        self.table_widget.setColumnWidth(self.table_widget.get_column_header().index("사진파일(경로)"), 350)

    def apply_common_input(self):
        grade = self.grade_input.text()
        name = self.name_input.text()
        install_date = self.install_date_input.text()
        survey_date = self.survey_date_input.text()        
        surveyor_position = self.surveyor_position_input.text()
        surveyor = self.surveyor_input.text()
        findings = self.findings_input.text()
        origin = self.origin_input.text()

        input_data = {'도선등급': grade, '도선명': name, '설치년월일': install_date, '조사년월일': survey_date, 
                      "조사자(직)": surveyor_position, "조사자(성명)": surveyor, "조사내용": findings, "원점": origin}
        
        for k, v in input_data.items():
            self.table_widget.set_column_value(k, v)
        
        self.status_message.setText('공통값 입력을 완료했습니다.')

    def get_image_folder(self):
        image_path = QFileDialog.getExistingDirectory(self, "Image Path")
        if image_path:
            self.image_folder = image_path

    def onRadioButtonToggled(self):
        radioButton = self.sender()
        if radioButton.isChecked():
            self.image_extension = radioButton.text()
            self.status_message.setText(f'{self.image_extension}를 확장자로 선택했습니다.')

    def apply_image_settings(self):
        if self.image_folder:
            self.table_widget.set_column_value("사진파일(경로)", self.image_folder)
        for row in range(self.table_widget.rowCount()):
            num = self.table_widget.item(row, 0).text()
            if self.image_edited_check.isChecked():
                self.png_radio.setChecked(True)
                self.table_widget.setCellItemAligned(row, self._headerindex("사진파일명"), ''.join([num, '_edit',self.image_extension]))
            else:
                self.table_widget.setCellItemAligned(row, self._headerindex("사진파일명"), ''.join([num, self.image_extension]))
        self.status_message.setText("사진정보를 갱신하였습니다.")

    def setLocation(self, kind_of_db):
        db_file = {'cif': 'Cif Files (*.cif)', 'shp': 'Esri Shp File (*.shp)'}
        jijuk, _ = QFileDialog.getOpenFileName(self,"Get Jijuk DB", "", f"{db_file.get(kind_of_db.lower())};;All Files (*)")
        try:
            if jijuk:
                if jijuk.lower().endswith(".cif"):
                    gdf = CifGeoDataFrame(jijuk).convert_to_geodataframe()
                if jijuk.lower().endswith(".shp"):
                    gdf = gpd.read_file(jijuk, encoding='euc-kr')
                
                if gdf is None: 
                    return
                
                not_found = []
                for i in range(self.table_widget.rowCount()):
                    location = find_attributes_containing_point(gdf, (float(self.table_widget.item(i, 2).text()), 
                                                                      float(self.table_widget.item(i, 1).text())),
                                                                      ["PNU", "JIBUNJIMOK", "DOHO"])
                    if not location is None:
                        pnu, jibun, dom = location.iloc[0, :]
                        self.table_widget.item(i, 6).setText(CifGeoDataFrame().getDistrictName(pnu))
                        self.table_widget.item(i, 7).setText(CifGeoDataFrame().pnu2jibun(pnu))
                        self.table_widget.item(i, 8).setText(self.dom_to_doho(dom)) 
                    else:
                        not_found.append(self.table_widget.item(i, 0).text())
                        self.table_widget.item(i, 6).setText("")
                        self.table_widget.item(i, 7).setText("")
                        self.table_widget.item(i, 8).setText("") 

                message = "주소검색을 마쳤습니다."
                if not_found:
                    message = f"주소검색을 마쳤습니다.\n미작성된 소재지 번호{str(not_found)}를 확인하세요."
    
                self.show_modal("success", parent=self.main_frame, title=" Searching is Done", description=message)    
                self.status_message.setText(f"{message}")

        except Exception as e:
            self.show_modal("error", parent=self.main_frame, title=" Unknown Error", description=f"{e}")
            self.status_message.setText(str(e))
    
    def dom_to_doho(self, dom):
        j_dom, _ = dom.split('/')
        return j_dom[1:-1]

    def saveProject(self):
        options = QFileDialog.Options()
        fileName, _ = QFileDialog.getSaveFileName(self, "Save project", "", "Pickle Files (*.pickle)", options=options)
        if fileName:
            try:
                self.save_table_to_pickle(self.table_widget, fileName)
                self.status_message.setText(f"파일이 성공적으로 저장되었습니다: {fileName}")
            except Exception as e:
                self.show_modal("error", parent=self.main_frame, title=" Cannot Save Pickle", description=f"파일 저장 중 오류 발생:\n{e}")

    def tablewidget_to_dataframe(self, table_widget: QTableWidget) -> pd.DataFrame:
        rows = table_widget.rowCount()
        columns = table_widget.columnCount()
        
        data = {}
        for column in range(columns):
            header = table_widget.horizontalHeaderItem(column).text()
            data[header] = []
            for row in range(rows):
                item = table_widget.item(row, column)
                data[header].append(item.text() if item else "")
        
        return pd.DataFrame(data)   

    def saveToExcel(self):
        fileName, _ = QFileDialog.getSaveFileName(self, caption="Save Excel File", dir=self.rtk_data_path, filter="Excel Files (*.xlsx)")
        if fileName:
            table_df = self.tablewidget_to_dataframe(self.table_widget)
            table_df.fillna("", inplace=True)
            # 서식에 대한 스타일 적용
            # # 구 성과표
            # border_settings =[{"rng": "A3:AF25","edges": ["all"], "border_style": "hair", "reset": True },  
            #                 {"rng": "A3:AF25","edges": ["outer"], "border_style": "thin",  "reset": False },
            #                 {"rng": "A3:A4","edges": ["inner_horizontal"], "border_style": None,  "reset": False },
            #                 {"rng": "A6:A7","edges": ["inner_horizontal"], "border_style": None, "reset": False }, 
            #                 {"rng": "A16:AF19","edges": ["inner_vertical"], "border_style": None, "reset": False },
            #                 {"rng": "A16:AF19","edges": ["inner_horizontal"], "border_style": None, "reset": False },
            #                 {"rng": "B8:AF9","edges": ["inner_horizontal"], "border_style": None, "reset": False },
            #                 {"rng": "B10:AF11","edges": ["inner_horizontal"], "border_style": None, "reset": False },
            #                 {"rng": "B12:AF13","edges": ["inner_horizontal"], "border_style": None, "reset": False },
            #                 {"rng": "N14:AF15","edges": ["inner_horizontal"], "border_style": None, "reset": False },
            #                 {"rng": "N3:S4","edges": ["inner_vertical"], "border_style": None, "reset": False }
            # ]
            # # 값 입력
            # mappings = [ {'fields': '점번호', 'address': 'B3', "callback":str_deco, 'kargs':{"postfix":"번"}},
            #             {'fields': '도선등급', 'address': 'G3', "callback":str_deco, 'kargs':{"postfix":"등"}},
            #             {'fields': '도선명', 'address': 'N3'},
            #             {'fields': '표지재질', 'address': 'Z3'},
            #             {'fields':['토지소재(동리)', '토지소재(지번)'], 'address': 'B5', 'callback': str_add, 'kargs':{'delim': ' ', 'postfix': "번지"}},  
            #             {'fields': '지적(임야)도', 'address': 'Z5', "callback":str_deco, 'kargs':{"postfix":"호"}},
            #             {'fields': '설치년월일', 'address': 'A8', 'callback': hangul_date},
            #             {'fields': 'X', 'address': 'B9', 'callback':osa},
            #             {'fields': 'Y', 'address': 'F9', 'callback':osa},
            #             {'fields': '경위도(B)', 'address': 'M9', 'callback':toBL},
            #             {'fields': '경위도(L)', 'address': 'W9', 'callback':toBL},
            #             {'fields': '원점', 'address': 'B14'},
            #             {'fields': '표고', 'address': 'N15'},
            #             {'fields': '조사년월일', 'address': 'A22', 'callback': hangul_date},
            #             {'fields': '조사자(직)', 'address': 'B22'},
            #             {'fields': '조사자(성명)', 'address': 'F22'},
            #             {'fields': '조사내용', 'address': 'L22'},
            #             {'fields': ['사진파일(경로)', '사진파일명'], 'address': "A17:AF19", 'callback': insert_image, 'kargs':{'keep_ratio': True}}
            # ]  

            # 신 성과표
            border_settings =[{"rng": "A6:L37","edges": ["all"], "border_style": "hair", "reset": True },  
                            {"rng": "A6:L37","edges": ["outer"], "border_style": "thin",  "reset": False },
                            {"rng": "A6:A7","edges": ["inner_horizontal"], "border_style": None, "reset": False }, 
                            {"rng": "A10:A11","edges": ["inner_horizontal"], "border_style": None,  "reset": False },
                            {"rng": "I12:L13","edges": ["inner_horizontal"], "border_style": None, "reset": False },
                            {"rng": "I14:L15","edges": ["inner_horizontal"], "border_style": None, "reset": False },
                            {"rng": "I16:L17","edges": ["inner_horizontal"], "border_style": None, "reset": False },
                            {"rng": "A22:G30","edges": ["inner_horizontal"], "border_style": None, "reset": False },
                            {"rng": "A22:G30","edges": ["inner_vertical"], "border_style": None, "reset": False },
                            {"rng": "H22:L30","edges": ["inner_horizontal"], "border_style": None, "reset": False },
                            {"rng": "H22:L30","edges": ["inner_vertical"], "border_style": None, "reset": False },
                            {"rng": "H31:L33","edges": ["inner_horizontal"], "border_style": None, "reset": False }
            ]
            # 값 입력
            mappings = [ {'fields': '점번호', 'address': 'D6', "callback":str_deco, 'kargs':{"postfix":"번", "allow_empty": False}},
                        {'fields': '도선등급', 'address': 'G6', "callback":str_deco, 'kargs':{"postfix":"등", "allow_empty": False}},
                        {'fields': '도선명', 'address': 'I6', "callback":str_deco, 'kargs':{"postfix":"도선", "allow_empty": False}},
                        {'fields': '표지재질', 'address': 'L6'},
                        {'fields':['토지소재(동리)', '토지소재(지번)'], 'address': 'B8', 'callback': str_add, 'kargs':{'delim': ' ', 'postfix': "번지"}},  
                        {'fields': '지적(임야)도', 'address': 'L8', "callback":str_deco, 'kargs':{"postfix":"호", "allow_empty": False}},
                        {'fields': '설치년월일', 'address': 'A12', 'callback': hangul_date},
                        {'fields': 'X', 'address': 'B12', 'callback':osa},
                        {'fields': 'Y', 'address': 'G12', 'callback':osa},
                        {'fields': '경위도(B)', 'address': 'I13', 'callback':toBL},
                        {'fields': '경위도(L)', 'address': 'K13', 'callback':toBL},
                        {'fields': '원점', 'address': 'B18'},
                        {'fields': '표고', 'address': 'K18'},
                        {'fields': '조사년월일', 'address': 'A34', 'callback': hangul_date},
                        {'fields': '조사자(직)', 'address': 'B34'},
                        {'fields': '조사자(성명)', 'address': 'F34'},
                        {'fields': '조사내용', 'address': 'H34'},
                        {'fields': ['사진파일(경로)', '위성사진'], 'address': "A22:G30", 'callback': insert_image, 'kargs':{'keep_ratio': False}},
                        {'fields': ['사진파일(경로)', '사진파일명'], 'address': "H22:L30", 'callback': insert_image, 'kargs':{'keep_ratio': False}}
            ] 

            try:
                # 실제 성과표(구) 작성
                temporary_path = '_기준점성과표.xlsx'
                # success = self.copy_resource_to_file(self.TEMPLATE, temporary_path)
                self.copy_resource_to_file(self.TEMPLATE2, temporary_path)
                  
                # 구 성과표
                # reporter = ReportFromDataframe(template=temporary_path, sheetname='서식', savefile=fileName, dataframe=table_df, 
                #                             max_row=26, border_settings=border_settings, mappings=mappings)
                # 
                reporter = ReportFromDataframe(template=temporary_path, savefile=fileName, dataframe=table_df, 
                                            max_row=38, border_settings=border_settings, mappings=mappings)
                reporter.report()
                self.status_message.setText(f"[기준점성과표]가 성공적으로 저장되었습니다: {fileName}")
                self.show_modal("success", parent=self.main_frame, title=" Saving Success", description=f"[기준점성과표]가 성공적으로 저장되었습니다.\n{fileName}")

            except FileExistsError as e:
                self.show_modal("error", parent=self.main_frame, title=" File Copy Failed", description=f"[기준점성과표] 파일 복사에 실패했습니다.")
                return

    def on_update_code(self):
        try:
            db = CodeGoKr().get_db()
            if not db is None:
                self.status_message.setText(f'Code 업데이트 완료...{db}')
                self.show_modal("success", parent=self.main_frame, title=" Update Code", description=f"Code 업데이트 완료...\n{db}")
        except Exception as e:
            self.show_modal("error", parent=self.main_frame, title=" Cannot Update Code", description=f"Code 업데이트 중 오류 발생:\n{e}")  


    def on_classify_image(self):
        dialog = DialogRenameImage()
        dialog.setObjectName("dialogrenameimage")
        dialog.setStyleSheet(self.get_stylesheet_from_resource(':resources/styles/dialogrenameimage.qss'))
        if not self.tr is None:
            dialog.tr = self.tr
        dialog.exec()

    def _headerindex(self, label:str, listofheaders = None) -> int:
        listofheaders = self.HEADER_LABELS if listofheaders is None else listofheaders
        return listofheaders.index(label)
    
    def get_stylesheet_from_resource(self, resource_path):
        qss_file = QFile(resource_path)
        if qss_file.open(QFile.ReadOnly | QFile.Text):  # 파일 열기
            stream = QTextStream(qss_file)  # QTextStream으로 파일 읽기
            stylesheet = stream.readAll()   # 모든 내용 읽어오기
            qss_file.close()  # 파일 닫기
            return stylesheet

    def show_settings_dialog(self):        
        dialog = Settings()
        dialog.setObjectName("settings")
        dialog.setStyleSheet(self.get_stylesheet_from_resource(':resources/styles/dialogrenameimage.qss'))
        
        if dialog.exec() == QDialog.Accepted:
            self.settings = dialog.settings
            self.reflect_settings()

    def reflect_settings(self):
        current_user = self.settings.get_current_user()
        if current_user is None:
            return
        users = self.settings.get_all_user()
        self.surveyor_name.clear()        
        self.surveyor_name.addItems(["팀 선택"] + users)
        if current_user in users:
            self.surveyor_name.setCurrentIndex(users.index(current_user) + 1)
            user_data = self.settings.get_section(current_user)
            self.surveyor_grade.setCurrentText(user_data["grade"])
            self.survey_license.setCurrentText(user_data["license"])
            self.antena_sn.setText(user_data["machine_serial"])
        else:
            self.surveyor_name.setCurrentIndex(0)

    def surveyor_changed(self):
        user = self.surveyor_name.currentText()
        if user != "팀 선택" and self.input_rtkdata_button.isChecked():
            self.settings.add_section("CUR_USER", {"user": self.surveyor_name.currentText()})
            user_data = self.settings.get_section(user)
            self.surveyor_grade.setCurrentText(user_data["grade"])
            self.survey_license.setCurrentText(user_data["license"])
            self.antena_sn.setText(user_data["machine_serial"])
            
 
if __name__ == "__main__":
    app=QApplication(sys.argv)
    ex = Splashscreen()
    sys.exit(app.exec())